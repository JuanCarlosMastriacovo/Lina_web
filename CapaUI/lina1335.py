from fastapi import APIRouter, Request, Query
from fastapi.responses import Response, HTMLResponse, RedirectResponse
from weasyprint import HTML
from jinja2 import Environment, FileSystemLoader
from pathlib import Path
from datetime import date, datetime
from io import BytesIO

import openpyxl

from CapaBRL.linabase import linabase
from CapaBRL.formatters import fmt_money
from CapaBRL.stock_brl import get_existencias_batch, get_ventas_periodo
from CapaBRL.txt_brl import generar_txt, col as txt_col
from CapaDAL.tablebase import get_table_model
from CapaBRL.config import APP_CONFIG
from CapaUI.xlsx_styles import (
    TITLE_FONT, SUBTITLE_FONT, HEADER_FONT, HEADER_FILL, HEADER_ALIGN,
    TOTAL_FONT, CURRENCY_FORMAT,
)

# ==================== CONSTANTES Y ROUTER ====================

router     = APIRouter()
PROG_CODE  = "LINA1335"
ROUTE_BASE = "/lina1335"

LinaArti          = get_table_model("linaarti")
ARTICLE_KEY_FIELD = LinaArti.get_business_key_field()

pdf_templates_dir = Path(__file__).parent.parent / "templates"
pdf_jinja_env     = Environment(loader=FileSystemLoader(str(pdf_templates_dir)))


# ==================== CLASE PRINCIPAL ====================

class Lina1335(linabase):
    """Módulo de ventas de artículos por período (LINA1335)."""
    pass


# ==================== FUNCIONES AUXILIARES ====================

def _get_filas(fecha_ini: date, fecha_fin: date) -> list:
    """
    Retorna filas [articodi, artidesc, vendidas, artiprec, valor, artipmpe, existencia]
    para TODOS los artículos (vendidas=0 si no hubo ventas en el período),
    ordenadas por valor DESC, luego articodi ASC.

    Existencia calculada al fecha_fin usando la misma lógica que lina1333/1334.
    """
    # list_all gestiona su propia conexión del pool (aplica filtro de empresa)
    arts = LinaArti.list_all(
        order_by=ARTICLE_KEY_FIELD,
        fields=["articodi", "artidesc", "artiprec", "artipmpe"],
    )
    if not arts:
        return []

    articodis = [str(a.get("articodi") or "").strip() for a in arts]
    venta_map = get_ventas_periodo(articodis, fecha_ini, fecha_fin)

    # Existencia al fecha_fin (cardex)
    exist_map = get_existencias_batch(articodis, fecha_fin)

    filas = []
    for a in arts:
        articodi = str(a.get("articodi") or "").strip()
        artidesc = str(a.get("artidesc") or "")
        artiprec = float(a.get("artiprec") or 0)
        artipmpe = int(a.get("artipmpe") or 0)
        vendidas = venta_map.get(articodi, 0)
        valor    = vendidas * artiprec
        exis     = exist_map.get(articodi, 0)
        filas.append([articodi, artidesc, vendidas, artiprec, valor, artipmpe, exis])

    # Orden: valor DESC, luego código ASC (igual que FoxPro IdxValo desc)
    filas.sort(key=lambda f: (-f[4], f[0]))
    return filas


# ==================== RUTAS ====================

@router.get("/", response_class=HTMLResponse)
async def lina1335_index(request: Request):
    Lina1335.set_prog_code(PROG_CODE)
    user = Lina1335.get_current_user(request)
    if not user:
        return RedirectResponse("/login")
    hoy = date.today().isoformat()
    return Lina1335.templates.TemplateResponse(
        "lina1335/seleccion.html",
        {"request": request, "error": None, "hoy": hoy},
    )


@router.get("/pdf")
async def lina1335_pdf(
    request: Request,
    fecha_ini: str = Query(default=""),
    fecha_fin: str = Query(default=""),
):
    Lina1335.set_prog_code(PROG_CODE)
    user = Lina1335.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    hoy = date.today()
    try:
        fi = date.fromisoformat(fecha_ini) if fecha_ini else hoy
    except ValueError:
        fi = hoy
    try:
        ff = date.fromisoformat(fecha_fin) if fecha_fin else hoy
    except ValueError:
        ff = hoy

    if fi > ff:
        return Lina1335.templates.TemplateResponse(
            "lina1335/seleccion.html",
            {"request": request, "error": "La fecha inicial no puede ser mayor que la final.", "hoy": hoy.isoformat()},
        )

    empr_code, empr_name = Lina1335.get_empr_info()
    filas = _get_filas(fi, ff)

    total_valor = sum(f[4] for f in filas)

    filas_pdf = [
        [f[0], f[1], f[2], fmt_money(f[3]), fmt_money(f[4]), f[5], f[6]]
        for f in filas
    ]

    template = pdf_jinja_env.get_template("lina1335/main.html")
    html_str = template.render(
        app_name        = APP_CONFIG.get("app_name", ""),
        app_description = APP_CONFIG.get("app_description", ""),
        prog_code   = PROG_CODE,
        empr_code   = empr_code,
        empr_name   = empr_name,
        usuario     = user,
        titulo      = "Ventas de Artículos por Período",
        fecha_ini   = fi.strftime("%d/%m/%Y"),
        fecha_fin   = ff.strftime("%d/%m/%Y"),
        fecha       = hoy.strftime("%d/%m/%Y"),
        hora        = datetime.now().strftime("%H:%M"),
        filas       = filas_pdf,
        total_valor = fmt_money(total_valor),
    )

    pdf = HTML(string=html_str).write_pdf()
    return Response(
        content    = pdf,
        media_type = "application/pdf",
        headers    = {"Content-Disposition": "inline; filename=ventas_articulos.pdf"},
    )


@router.get("/xlsx")
async def lina1335_xlsx(
    request: Request,
    fecha_ini: str = Query(default=""),
    fecha_fin: str = Query(default=""),
):
    Lina1335.set_prog_code(PROG_CODE)
    user = Lina1335.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    hoy = date.today()
    try:
        fi = date.fromisoformat(fecha_ini) if fecha_ini else hoy
    except ValueError:
        fi = hoy
    try:
        ff = date.fromisoformat(fecha_fin) if fecha_fin else hoy
    except ValueError:
        ff = hoy

    if fi > ff:
        return Lina1335.templates.TemplateResponse(
            "lina1335/seleccion.html",
            {"request": request, "error": "La fecha inicial no puede ser mayor que la final.", "hoy": hoy.isoformat()},
        )

    empr_code, empr_name = Lina1335.get_empr_info()
    filas = _get_filas(fi, ff)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas por Período"

    title_font    = TITLE_FONT
    subtitle_font = SUBTITLE_FONT
    header_font   = HEADER_FONT
    header_fill   = HEADER_FILL
    header_align  = HEADER_ALIGN
    total_font    = TOTAL_FONT

    ncols    = 7
    last_col = openpyxl.utils.get_column_letter(ncols)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"]      = "Ventas de Artículos por Período"
    ws["A1"].font = title_font

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"]      = f"{APP_CONFIG.get('app_name', '')} — {empr_code} {empr_name}"
    ws["A2"].font = subtitle_font

    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"]      = (f"Desde {fi.strftime('%d/%m/%Y')} hasta {ff.strftime('%d/%m/%Y')} — "
                     f"Usuario: {user} — "
                     f"{hoy.strftime('%d/%m/%Y')} {datetime.now().strftime('%H:%M')}")
    ws["A3"].font = subtitle_font

    ws.append([])   # fila 4 vacía

    headers = ["Código", "Descripción", "Vendidas", "Precio", "Valor Vta.", "PMP", "Existencia"]
    ws.append(headers)   # fila 5
    for cell in ws[5]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    DATA_START = 6
    for i, fila in enumerate(filas):
        row_num = DATA_START + i
        # A=articodi B=artidesc C=vendidas D=artiprec E=valor F=artipmpe G=exis
        ws.append([fila[0], fila[1], fila[2], fila[3], None, fila[5], fila[6]])
        ws.cell(row=row_num, column=4).number_format = CURRENCY_FORMAT
        ws.cell(row=row_num, column=5).value         = f"=C{row_num}*D{row_num}"
        ws.cell(row=row_num, column=5).number_format = CURRENCY_FORMAT

    total_row = DATA_START + len(filas)
    last_data = total_row - 1
    ws.append(["", "TOTAL", "", "", f"=SUM(E{DATA_START}:E{last_data})", "", ""])
    for cell in ws[total_row]:
        cell.font = total_font
    ws.cell(row=total_row, column=5).number_format = CURRENCY_FORMAT

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 12

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        content    = buffer.read(),
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers    = {"Content-Disposition": "attachment; filename=ventas_articulos.xlsx"},
    )


_COLUMNAS_TXT = [
    txt_col("Código",     9,  "L"),
    txt_col("Descripción",34, "L"),
    txt_col("Vendidas",   9,  "R"),
    txt_col("Precio",     12, "R"),
    txt_col("Valor Vta.", 12, "R"),
    txt_col("PMP",        5,  "R"),
    txt_col("Exis.",      7,  "R"),
]


@router.get("/txt")
async def lina1335_txt(
    request: Request,
    fecha_ini: str = Query(default=""),
    fecha_fin: str = Query(default=""),
):
    Lina1335.set_prog_code(PROG_CODE)
    user = Lina1335.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    hoy = date.today()
    try:
        fi = date.fromisoformat(fecha_ini) if fecha_ini else hoy
    except ValueError:
        fi = hoy
    try:
        ff = date.fromisoformat(fecha_fin) if fecha_fin else hoy
    except ValueError:
        ff = hoy

    if fi > ff:
        return Lina1335.templates.TemplateResponse(
            "lina1335/seleccion.html",
            {"request": request, "error": "La fecha inicial no puede ser mayor que la final.", "hoy": hoy.isoformat()},
        )

    empr_code, empr_name = Lina1335.get_empr_info()
    filas = _get_filas(fi, ff)

    total_valor = sum(f[4] for f in filas)

    filas_txt = [
        [f[0], f[1], str(f[2]), fmt_money(f[3]), fmt_money(f[4]), str(f[5]), str(f[6])]
        for f in filas
    ]
    totales_txt = ["", "TOTAL", "", "", fmt_money(total_valor), "", ""]

    txt = generar_txt(
        titulo    = "Ventas de Artículos por Período",
        subtitulo = f"Desde {fi.strftime('%d/%m/%Y')} hasta {ff.strftime('%d/%m/%Y')}",
        columnas  = _COLUMNAS_TXT,
        filas     = filas_txt,
        app_name  = APP_CONFIG.get("app_name", ""),
        empr_code = empr_code,
        empr_name = empr_name,
        usuario   = user,
        fecha     = hoy.strftime("%d/%m/%Y"),
        hora      = datetime.now().strftime("%H:%M"),
        totales   = totales_txt,
    )

    return Response(
        content    = txt.encode("utf-8"),
        media_type = "text/plain; charset=utf-8",
        headers    = {"Content-Disposition": "inline; filename=ventas_articulos.txt"},
    )
