from fastapi import APIRouter, Request
from fastapi.responses import Response, HTMLResponse, RedirectResponse
from weasyprint import HTML
from jinja2 import Environment, FileSystemLoader
from pathlib import Path
from datetime import date, datetime
from io import BytesIO
from itertools import groupby as _groupby

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from CapaBRL.linabase import linabase
from CapaBRL.listaprecios_brl import aplicar_reglas
from CapaDAL.tablebase import get_table_model
from CapaDAL.dataconn import ctx_empr
from CapaDAL.config import APP_CONFIG

# ==================== CONSTANTES Y ROUTER ====================

router     = APIRouter()
PROG_CODE  = "LINA1332"
ROUTE_BASE = "/lina1332"

LinaArti          = get_table_model("linaarti")
LinaArtr          = get_table_model("linaartr")
LinaEmpr          = get_table_model("linaempr")
ARTICLE_KEY_FIELD = LinaArti.get_business_key_field()
ARTR_KEY_FIELD    = LinaArtr.get_business_key_field()
EMPR_CODE_FIELD   = LinaEmpr.require_column("emprcodi")
EMPR_NAME_FIELD   = LinaEmpr.require_column("emprname")

pdf_templates_dir = Path(__file__).parent.parent / "templates"
pdf_jinja_env     = Environment(loader=FileSystemLoader(str(pdf_templates_dir)))


# ==================== CLASE PRINCIPAL ====================

class Lina1332(linabase):
    """Módulo de lista de precios de artículos (LINA1332)."""
    pass


# ==================== FUNCIONES AUXILIARES ====================

def _get_empr_info():
    empr_code = ctx_empr.get() or "01"
    empr_rec  = LinaEmpr.row_get({EMPR_CODE_FIELD: empr_code})
    empr_name = str(empr_rec.get(EMPR_NAME_FIELD) or "").strip() if empr_rec else ""
    return empr_code, empr_name


def _fmt_price(val: float) -> str:
    """Precio con punto de miles y coma decimal: 1.234,56"""
    return f"{val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _get_grupos(conn):
    """
    Devuelve lista de grupos ordenados por artrcodi+articodi.
    Cada grupo: {"artrcodi": str, "artrdesc": str, "filas": [[cod, desc, prec_float], ...]}
    """
    arts = LinaArti.list_all(
        order_by=ARTICLE_KEY_FIELD,
        fields=["articodi", "artidesc", "artrcodi", "artiprec"],
        conn=conn,
    )

    # Lookup de descripciones de rubro
    rubros = {}
    try:
        rubro_field_desc = LinaArtr.get_selector_fields()[1]
        for r in LinaArtr.list_all(fields=[ARTR_KEY_FIELD, rubro_field_desc], conn=conn):
            rubros[str(r.get(ARTR_KEY_FIELD) or "").strip()] = str(r.get(rubro_field_desc) or "").strip()
    except Exception:
        pass

    # Ordenar por (artrcodi, articodi) en Python
    sorted_arts = sorted(arts, key=lambda a: (
        str(a.get("artrcodi") or "").strip(),
        str(a.get("articodi") or "").strip(),
    ))

    grupos = []
    for artrcodi_val, group_iter in _groupby(
        sorted_arts, key=lambda a: str(a.get("artrcodi") or "").strip()
    ):
        artrdesc = rubros.get(artrcodi_val, "")
        filas = [
            [
                str(a.get("articodi") or "").strip(),
                str(a.get("artidesc") or ""),
                float(a.get("artiprec") or 0),
            ]
            for a in group_iter
        ]
        filas = aplicar_reglas(artrcodi_val, filas)
        if filas:
            grupos.append({"artrcodi": artrcodi_val, "artrdesc": artrdesc, "filas": filas})

    return grupos


# ==================== RUTAS ====================

@router.get("/", response_class=HTMLResponse)
async def lina1332_index(request: Request):
    Lina1332.set_prog_code(PROG_CODE)
    user = Lina1332.get_current_user(request)
    if not user:
        return RedirectResponse("/login")
    return Lina1332.templates.TemplateResponse(
        "lina1332/seleccion.html",
        {"request": request, "error": None},
    )


@router.get("/pdf")
async def lina1332_pdf(request: Request):
    """Genera el PDF de la lista de precios."""
    Lina1332.set_prog_code(PROG_CODE)
    user = Lina1332.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    empr_code, empr_name = _get_empr_info()
    conn   = Lina1332.get_task_conn(request, readonly=True)
    grupos = _get_grupos(conn)

    # Formatear precios para PDF
    for g in grupos:
        for fila in g["filas"]:
            fila[2] = _fmt_price(fila[2])

    total_articulos = sum(len(g["filas"]) for g in grupos)

    template = pdf_jinja_env.get_template("lina1332/main.html")
    html_str = template.render(
        app_name        = APP_CONFIG.get("app_name", ""),
        app_description = APP_CONFIG.get("app_description", ""),
        prog_code = PROG_CODE,
        empr_code = empr_code,
        empr_name = empr_name,
        usuario   = user,
        titulo    = "Lista de Precios",
        fecha     = date.today().strftime("%d/%m/%Y"),
        hora      = datetime.now().strftime("%H:%M"),
        grupos    = grupos,
        total_articulos = total_articulos,
    )

    pdf = HTML(string=html_str).write_pdf()
    return Response(
        content    = pdf,
        media_type = "application/pdf",
        headers    = {"Content-Disposition": "inline; filename=lista_precios.pdf"},
    )


@router.get("/xlsx")
async def lina1332_xlsx(request: Request):
    """Genera el XLSX de la lista de precios."""
    Lina1332.set_prog_code(PROG_CODE)
    user = Lina1332.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    empr_code, empr_name = _get_empr_info()
    conn   = Lina1332.get_task_conn(request, readonly=True)
    grupos = _get_grupos(conn)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Precios"

    title_font    = Font(bold=True, size=12)
    subtitle_font = Font(size=8, color="555555")
    header_font   = Font(bold=True, color="FFFFFF", size=9)
    header_fill   = PatternFill("solid", fgColor="4472C4")
    header_align  = Alignment(horizontal="left", vertical="center")
    grupo_font    = Font(bold=True, size=9, color="1F3864")
    grupo_fill    = PatternFill("solid", fgColor="BDD7EE")
    price_align   = Alignment(horizontal="right")

    PRICE_FMT = '#,##0.00;-#,##0.00'

    ws.merge_cells("A1:C1")
    ws["A1"]      = "Lista de Precios"
    ws["A1"].font = title_font

    ws.merge_cells("A2:C2")
    ws["A2"]      = f"{APP_CONFIG.get('app_name', '')} — {empr_code} {empr_name}"
    ws["A2"].font = subtitle_font

    ws.merge_cells("A3:C3")
    ws["A3"]      = f"Usuario: {user} — {date.today().strftime('%d/%m/%Y')} {datetime.now().strftime('%H:%M')}"
    ws["A3"].font = subtitle_font

    ws.append([])   # fila 4 vacía

    ws.append(["Código", "Descripción", "Precio"])
    for cell in ws[5]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    for idx, grupo in enumerate(grupos):
        if idx > 0:
            ws.append([])   # fila en blanco entre grupos

        titulo_rubro = grupo["artrcodi"]
        if grupo["artrdesc"]:
            titulo_rubro += " " + grupo["artrdesc"]

        ws.append([titulo_rubro, None, None])
        row_num = ws.max_row
        ws.merge_cells(f"A{row_num}:C{row_num}")
        for cell in ws[row_num]:
            cell.font  = grupo_font
            cell.fill  = grupo_fill
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="left")

        for fila in grupo["filas"]:
            ws.append(fila)
            data_row = ws.max_row
            ws.cell(row=data_row, column=3).number_format = PRICE_FMT
            ws.cell(row=data_row, column=3).alignment     = price_align

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 14

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        content    = buffer.read(),
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers    = {"Content-Disposition": "attachment; filename=lista_precios.xlsx"},
    )
