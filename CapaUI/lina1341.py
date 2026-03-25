from fastapi import APIRouter, Request, Query
from fastapi.responses import Response, HTMLResponse, RedirectResponse
from weasyprint import HTML
from jinja2 import Environment, FileSystemLoader
from pathlib import Path
from datetime import date, datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from CapaBRL.linabase import linabase
from CapaBRL.cardex_brl import get_cardex
from CapaDAL.config import APP_CONFIG
from CapaUI.xlsx_styles import (
    TITLE_FONT, SUBTITLE_FONT, HEADER_FONT, HEADER_FILL, HEADER_ALIGN,
)

# ==================== CONSTANTES Y ROUTER ====================

router     = APIRouter()
PROG_CODE  = "LINA1341"
ROUTE_BASE = "/lina1341"

pdf_templates_dir = Path(__file__).parent.parent / "templates"
pdf_jinja_env     = Environment(loader=FileSystemLoader(str(pdf_templates_dir)))

NUM_ALIGN = Alignment(horizontal="right")
APE_FONT  = Font(name="Arial", bold=True, size=9)
APE_FILL  = PatternFill("solid", fgColor="D9E2F3")


# ==================== CLASE PRINCIPAL ====================

class Lina1341(linabase):
    """Módulo de Ficha de Cardex (LINA1341)."""
    pass


# ==================== FUNCIONES AUXILIARES ====================

def _parse_params(articodi_raw: str, fecha_desde_raw: str, fecha_hasta_raw: str):
    """Devuelve (articodi, fd, ff, error_str). error_str es None si todo OK."""
    articodi = articodi_raw.strip().upper()
    if not articodi:
        return None, None, None, "Debe ingresar un código de artículo."

    hoy = date.today()
    fd  = None
    try:
        if fecha_desde_raw:
            fd = date.fromisoformat(fecha_desde_raw)
    except ValueError:
        pass
    try:
        ff = date.fromisoformat(fecha_hasta_raw) if fecha_hasta_raw else hoy
    except ValueError:
        ff = hoy

    if fd and fd > ff:
        return None, None, None, "La fecha inicial no puede ser mayor que la final."

    return articodi, fd, ff, None


def _build_subtitulo(art_info: dict, fd, ff: date) -> str:
    desde = fd.strftime("%d/%m/%Y") if fd else "inicio"
    return (f"{art_info['articodi']} — {art_info['artidesc']} — "
            f"Período: {desde} al {ff.strftime('%d/%m/%Y')}")


# ==================== RUTAS ====================

@router.get("/", response_class=HTMLResponse)
async def lina1341_index(request: Request):
    Lina1341.set_prog_code(PROG_CODE)
    user = Lina1341.get_current_user(request)
    if not user:
        return RedirectResponse("/login")
    return Lina1341.templates.TemplateResponse(
        "lina1341/seleccion.html",
        {"request": request, "error": None, "hoy": date.today().isoformat()},
    )


@router.get("/pdf")
async def lina1341_pdf(
    request: Request,
    articodi:    str = Query(default=""),
    fecha_desde: str = Query(default=""),
    fecha_hasta: str = Query(default=""),
):
    Lina1341.set_prog_code(PROG_CODE)
    user = Lina1341.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    articodi_p, fd, ff, err = _parse_params(articodi, fecha_desde, fecha_hasta)
    if err:
        return Lina1341.templates.TemplateResponse(
            "lina1341/seleccion.html",
            {"request": request, "error": err, "hoy": date.today().isoformat()},
        )

    empr_code, empr_name = Lina1341.get_empr_info()
    art_info, filas = get_cardex(articodi_p, fd, ff)

    if art_info is None:
        return Lina1341.templates.TemplateResponse(
            "lina1341/seleccion.html",
            {"request": request, "error": f"Artículo '{articodi_p}' no encontrado.",
             "hoy": date.today().isoformat()},
        )

    template = pdf_jinja_env.get_template("lina1341/main.html")
    html_str = template.render(
        app_name        = APP_CONFIG.get("app_name", ""),
        app_description = APP_CONFIG.get("app_description", ""),
        prog_code   = PROG_CODE,
        empr_code   = empr_code,
        empr_name   = empr_name,
        usuario     = user,
        fecha       = date.today().strftime("%d/%m/%Y"),
        hora        = datetime.now().strftime("%H:%M"),
        art_info    = art_info,
        fecha_desde = fd.strftime("%d/%m/%Y") if fd else "inicio",
        fecha_hasta = ff.strftime("%d/%m/%Y"),
        filas       = filas,
    )

    pdf = HTML(string=html_str).write_pdf()
    return Response(
        content    = pdf,
        media_type = "application/pdf",
        headers    = {"Content-Disposition": f"inline; filename=cardex_{articodi_p}.pdf"},
    )


@router.get("/xlsx")
async def lina1341_xlsx(
    request: Request,
    articodi:    str = Query(default=""),
    fecha_desde: str = Query(default=""),
    fecha_hasta: str = Query(default=""),
):
    Lina1341.set_prog_code(PROG_CODE)
    user = Lina1341.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    articodi_p, fd, ff, err = _parse_params(articodi, fecha_desde, fecha_hasta)
    if err:
        return Lina1341.templates.TemplateResponse(
            "lina1341/seleccion.html",
            {"request": request, "error": err, "hoy": date.today().isoformat()},
        )

    empr_code, empr_name = Lina1341.get_empr_info()
    art_info, filas = get_cardex(articodi_p, fd, ff)

    if art_info is None:
        return Lina1341.templates.TemplateResponse(
            "lina1341/seleccion.html",
            {"request": request, "error": f"Artículo '{articodi_p}' no encontrado.",
             "hoy": date.today().isoformat()},
        )

    subtitulo = _build_subtitulo(art_info, fd, ff)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cardex"

    # Bloque de título
    ws.merge_cells("A1:E1")
    ws["A1"]      = "Ficha de Cardex"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:E2")
    ws["A2"]      = f"{APP_CONFIG.get('app_name', '')} — {empr_code} {empr_name}"
    ws["A2"].font = SUBTITLE_FONT

    ws.merge_cells("A3:E3")
    ws["A3"]      = (f"{subtitulo}  |  PMP: {art_info['artipmpe']}  |  "
                     f"Precio: {art_info['artiprec']}  |  "
                     f"Usuario: {user} — "
                     f"{date.today().strftime('%d/%m/%Y')} {datetime.now().strftime('%H:%M')}")
    ws["A3"].font = SUBTITLE_FONT

    ws.append([])

    # Cabecera de columnas
    ws.append(["Fecha", "Concepto", "Salida", "Entrada", "Exist."])
    r = ws.max_row
    for cell in ws[r]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    for col in (3, 4, 5):
        ws.cell(r, col).alignment = NUM_ALIGN

    # Filas de datos
    for i, fila in enumerate(filas):
        sal  = int(fila[2]) if fila[2] else None
        ent  = int(fila[3]) if fila[3] else None
        sald = int(fila[4]) if fila[4] else 0
        ws.append([fila[0], fila[1], sal, ent, sald])
        r = ws.max_row
        for col in (3, 4, 5):
            ws.cell(r, col).alignment = NUM_ALIGN
        if i == 0:                    # fila de apertura
            for cell in ws[r]:
                cell.font = APE_FONT
                cell.fill = APE_FILL

    # Pie
    movs = max(len(filas) - 1, 0)
    ws.append([f"{movs} movimiento{'s' if movs != 1 else ''}"])
    ws.cell(ws.max_row, 1).font = SUBTITLE_FONT

    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(
        content    = buffer.read(),
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers    = {"Content-Disposition": f"attachment; filename=cardex_{articodi_p}.xlsx"},
    )


@router.get("/txt")
async def lina1341_txt(
    request: Request,
    articodi:    str = Query(default=""),
    fecha_desde: str = Query(default=""),
    fecha_hasta: str = Query(default=""),
):
    Lina1341.set_prog_code(PROG_CODE)
    user = Lina1341.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    articodi_p, fd, ff, err = _parse_params(articodi, fecha_desde, fecha_hasta)
    if err:
        return Lina1341.templates.TemplateResponse(
            "lina1341/seleccion.html",
            {"request": request, "error": err, "hoy": date.today().isoformat()},
        )

    empr_code, empr_name = Lina1341.get_empr_info()
    art_info, filas = get_cardex(articodi_p, fd, ff)

    if art_info is None:
        return Lina1341.templates.TemplateResponse(
            "lina1341/seleccion.html",
            {"request": request, "error": f"Artículo '{articodi_p}' no encontrado.",
             "hoy": date.today().isoformat()},
        )

    desde_str = fd.strftime("%d/%m/%Y") if fd else "inicio"
    SEP = "-" * 66
    HDR = f"{'Fecha':<12}  {'Concepto':<22}  {'Salida':>8}  {'Entrada':>8}  {'Exist.':>8}"

    lines = [
        "FICHA DE CARDEX",
        f"{art_info['articodi']} — {art_info['artidesc']}",
        f"Período: {desde_str} al {ff.strftime('%d/%m/%Y')}  |  PMP: {art_info['artipmpe']}  |  Precio: {art_info['artiprec']}",
        f"{APP_CONFIG.get('app_name', '')} — {empr_code} {empr_name}",
        f"Usuario: {user} — {date.today().strftime('%d/%m/%Y')} {datetime.now().strftime('%H:%M')}",
        "",
        HDR,
        SEP,
    ]

    for fila in filas:
        sal  = fila[2].rjust(8) if fila[2] else " " * 8
        ent  = fila[3].rjust(8) if fila[3] else " " * 8
        sald = fila[4].rjust(8)
        lines.append(f"{fila[0]:<12}  {fila[1]:<22}  {sal}  {ent}  {sald}")

    movs = max(len(filas) - 1, 0)
    lines += [SEP, f"{movs} movimiento{'s' if movs != 1 else ''}."]

    txt = "\r\n".join(lines)
    return Response(
        content    = txt.encode("utf-8"),
        media_type = "text/plain; charset=utf-8",
        headers    = {"Content-Disposition": f"inline; filename=cardex_{articodi_p}.txt"},
    )
