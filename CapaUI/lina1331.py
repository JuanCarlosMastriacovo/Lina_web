from fastapi import APIRouter, Request, Query
from fastapi.responses import Response, HTMLResponse, RedirectResponse
from weasyprint import HTML
from jinja2 import Environment, FileSystemLoader
from pathlib import Path
from datetime import date, datetime
from io import BytesIO

import openpyxl

from CapaBRL.linabase import linabase
from CapaBRL.formatters import fmt_money
from CapaDAL.tablebase import get_table_model
from CapaDAL.config import APP_CONFIG
from CapaUI.xlsx_styles import (
    TITLE_FONT, SUBTITLE_FONT, HEADER_FONT, HEADER_FILL, HEADER_ALIGN,
    CURRENCY_FORMAT, DATE_FMT,
)

# ==================== CONSTANTES Y ROUTER ====================

router     = APIRouter()
PROG_CODE  = "LINA1331"
ROUTE_BASE = "/lina1331"

LinaArti          = get_table_model("linaarti")
ARTICLE_KEY_FIELD = LinaArti.get_business_key_field()   # articodi

pdf_templates_dir = Path(__file__).parent.parent / "templates"
pdf_jinja_env     = Environment(loader=FileSystemLoader(str(pdf_templates_dir)))

# Índices de columna dentro de cada fila (constantes para ambos órdenes)
# fila[0..2] = campos variables (cod/rub según orden)
# fila[3]  = PMP       (entero)
# fila[4]  = Precio    (float)
# fila[5]  = Ex.Ant.   (entero)
# fila[6]  = Fe.Ex.Ant (date | None)
# fila[7]  = Últ.Costo (float)
# fila[8]  = Fe.Últ.C. (date | None)
# fila[9]  = Últ.C.Cant (entero)
_PRICE_IDX = (4, 7)
_DATE_IDX  = (6, 8)

_HEADERS_CODIGO = ["Código", "Descripción", "Rubro", "PMP", "Precio", "Ex.Ant.", "Fe.Ex.Ant.", "Últ.Costo", "Fe.Últ.C.", "Últ.C.Cant."]
_HEADERS_RUBRO  = ["Rubro", "Código", "Descripción", "PMP", "Precio", "Ex.Ant.", "Fe.Ex.Ant.", "Últ.Costo", "Fe.Últ.C.", "Últ.C.Cant."]

_COL_WIDTHS_CODIGO = [11, 38, 8, 8, 12, 8, 12, 12, 12, 12]
_COL_WIDTHS_RUBRO  = [8,  11, 38, 8, 12, 8, 12, 12, 12, 12]


# ==================== CLASE PRINCIPAL ====================

class Lina1331(linabase):
    """Módulo de listado detallado de artículos (LINA1331)."""
    pass


# ==================== FUNCIONES AUXILIARES ====================

def _parse_date(val):
    """Convierte un valor de BD a date, o None si no es válido."""
    if val is None:
        return None
    if isinstance(val, date):
        return val
    s = str(val).strip()[:10]
    if not s or s.lower() == "none":
        return None
    try:
        return date.fromisoformat(s)
    except (ValueError, TypeError):
        return None


def _get_articulos(conn, desde: str, hasta: str, orden: str = "codigo"):
    """
    Devuelve filas con valores raw:
      float para Precio y Últ.Costo, date|None para fechas.
    Los llamadores formatean según destino (PDF → strings, XLSX → tipos nativos).
    """
    arts = LinaArti.list_all(
        order_by=ARTICLE_KEY_FIELD,
        fields=[
            "articodi", "artidesc", "artrcodi",
            "artipmpe", "artiprec",
            "artiexan", "artiexfe",
            "artiucco", "artiucfe", "artiucca",
        ],
        conn=conn,
    )

    filas = []
    for a in arts:
        cod = str(a.get("articodi") or "").strip()
        if desde and cod < desde:
            continue
        if hasta and cod > hasta:
            continue

        artrcodi = str(a.get("artrcodi") or "").strip()
        prec     = float(a.get("artiprec") or 0)
        ucco     = float(a.get("artiucco") or 0)
        exfe     = _parse_date(a.get("artiexfe"))
        ucfe     = _parse_date(a.get("artiucfe"))

        tail = [
            str(a.get("artipmpe") or 0),   # [3] PMP
            prec,                            # [4] Precio     ← float
            str(a.get("artiexan") or 0),    # [5] Ex.Ant.
            exfe,                            # [6] Fe.Ex.Ant. ← date|None
            ucco,                            # [7] Últ.Costo  ← float
            ucfe,                            # [8] Fe.Últ.C.  ← date|None
            str(a.get("artiucca") or 0),    # [9] Últ.C.Cant.
        ]

        if orden == "rubro":
            filas.append([artrcodi, cod, str(a.get("artidesc") or "")] + tail)
        else:
            filas.append([cod, str(a.get("artidesc") or ""), artrcodi] + tail)

    if orden == "rubro":
        filas.sort(key=lambda r: (r[0], r[1]))

    return filas


def _filas_for_pdf(filas):
    """Convierte los valores raw a strings para el template PDF."""
    result = []
    for f in filas:
        row = list(f)
        row[4] = fmt_money(row[4])
        row[6] = row[6].strftime("%d/%m/%Y") if row[6] else ""
        row[7] = fmt_money(row[7])
        row[8] = row[8].strftime("%d/%m/%Y") if row[8] else ""
        result.append(row)
    return result


# ==================== RUTAS ====================

@router.get("/", response_class=HTMLResponse)
async def lina1331_index(request: Request):
    Lina1331.set_prog_code(PROG_CODE)
    user = Lina1331.get_current_user(request)
    if not user:
        return RedirectResponse("/login")
    return Lina1331.templates.TemplateResponse(
        "lina1331/seleccion.html",
        {"request": request, "error": None},
    )


@router.get("/pdf")
async def lina1331_pdf(
    request: Request,
    desde: str = Query(default=""),
    hasta: str = Query(default="ZZZZZZZZZ"),
    orden: str = Query(default="codigo"),
):
    """Genera el PDF del listado detallado de artículos."""
    Lina1331.set_prog_code(PROG_CODE)
    user = Lina1331.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    desde = desde.strip().upper()
    hasta = hasta.strip().upper()
    orden = orden if orden in ("codigo", "rubro") else "codigo"

    if desde and hasta and desde > hasta:
        return Lina1331.templates.TemplateResponse(
            "lina1331/seleccion.html",
            {"request": request, "error": "El valor 'Desde' no puede ser mayor que 'Hasta'."},
        )

    empr_code, empr_name = Lina1331.get_empr_info()
    conn  = Lina1331.get_task_conn(request, readonly=True)
    filas = _filas_for_pdf(_get_articulos(conn, desde, hasta, orden))

    subtitulo = f"Artículos {desde or '(inicio)'} al {hasta or '(fin)'}"

    template = pdf_jinja_env.get_template("lina1331/main.html")
    html_str = template.render(
        app_name        = APP_CONFIG.get("app_name", ""),
        app_description = APP_CONFIG.get("app_description", ""),
        prog_code = PROG_CODE,
        empr_code = empr_code,
        empr_name = empr_name,
        usuario   = user,
        titulo    = "Listado Detallado de Artículos",
        subtitulo = subtitulo,
        fecha     = date.today().strftime("%d/%m/%Y"),
        hora      = datetime.now().strftime("%H:%M"),
        orden     = orden,
        filas     = filas,
    )

    pdf = HTML(string=html_str).write_pdf()
    return Response(
        content    = pdf,
        media_type = "application/pdf",
        headers    = {"Content-Disposition": "inline; filename=articulos_detallado.pdf"},
    )


@router.get("/xlsx")
async def lina1331_xlsx(
    request: Request,
    desde: str = Query(default=""),
    hasta: str = Query(default="ZZZZZZZZZ"),
    orden: str = Query(default="codigo"),
):
    """Genera el XLSX del listado detallado de artículos."""
    Lina1331.set_prog_code(PROG_CODE)
    user = Lina1331.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    desde = desde.strip().upper()
    hasta = hasta.strip().upper()
    orden = orden if orden in ("codigo", "rubro") else "codigo"

    if desde and hasta and desde > hasta:
        return Lina1331.templates.TemplateResponse(
            "lina1331/seleccion.html",
            {"request": request, "error": "El valor 'Desde' no puede ser mayor que 'Hasta'."},
        )

    empr_code, empr_name = Lina1331.get_empr_info()
    conn  = Lina1331.get_task_conn(request, readonly=True)
    filas = _get_articulos(conn, desde, hasta, orden)   # raw values

    subtitulo = f"Artículos {desde or '(inicio)'} al {hasta or '(fin)'}"
    headers   = _HEADERS_RUBRO if orden == "rubro" else _HEADERS_CODIGO
    widths    = _COL_WIDTHS_RUBRO if orden == "rubro" else _COL_WIDTHS_CODIGO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Artículos"

    header_font   = HEADER_FONT
    header_fill   = HEADER_FILL
    header_align  = HEADER_ALIGN
    title_font    = TITLE_FONT
    subtitle_font = SUBTITLE_FONT

    ncols    = 10
    last_col = openpyxl.utils.get_column_letter(ncols)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"]      = "Listado Detallado de Artículos"
    ws["A1"].font = title_font

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"]      = f"{APP_CONFIG.get('app_name', '')} — {empr_code} {empr_name}"
    ws["A2"].font = subtitle_font

    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"]      = f"{subtitulo} — Usuario: {user} — {date.today().strftime('%d/%m/%Y')} {datetime.now().strftime('%H:%M')}"
    ws["A3"].font = subtitle_font

    ws.append([])   # fila 4 vacía

    ws.append(headers)   # fila 5 = encabezados
    for cell in ws[5]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    # Filas de datos (empiezan en fila 6)
    for row_num, fila in enumerate(filas, start=6):
        ws.append(fila)
        for col_idx in _PRICE_IDX:
            ws.cell(row=row_num, column=col_idx + 1).number_format = CURRENCY_FORMAT
        for col_idx in _DATE_IDX:
            cell = ws.cell(row=row_num, column=col_idx + 1)
            if cell.value is not None:
                cell.number_format = DATE_FMT

    for idx, w in enumerate(widths):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = w

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        content    = buffer.read(),
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers    = {"Content-Disposition": "attachment; filename=articulos_detallado.xlsx"},
    )
