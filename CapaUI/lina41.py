from fastapi import APIRouter, Request, Query
from fastapi.responses import Response, HTMLResponse, RedirectResponse
from weasyprint import HTML
from jinja2 import Environment, FileSystemLoader
from pathlib import Path
from datetime import date, datetime, timedelta
from io import BytesIO

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

from CapaBRL.linabase import linabase
from CapaBRL.formatters import fmt_money
from CapaDAL.dataconn import sess_conns, ctx_empr, ctx_date
from CapaBRL.config import APP_CONFIG
from CapaBRL.config import DEFAULT_EMPR_CODE
from CapaUI.xlsx_styles import (
    TITLE_FONT, SUBTITLE_FONT, HEADER_FONT, HEADER_FILL, HEADER_ALIGN, CURRENCY_FORMAT,
)

# ==================== CONSTANTES Y ROUTER ====================

router     = APIRouter()
PROG_CODE  = "LINA41"
ROUTE_BASE = "/lina41"

pdf_templates_dir = Path(__file__).parent.parent / "templates"
pdf_jinja_env     = Environment(loader=FileSystemLoader(str(pdf_templates_dir)))


# ==================== CLASE PRINCIPAL ====================

class Lina41(linabase):
    """Listado de Movimientos de Caja (LINA41)."""
    pass


# ==================== FUNCIONES AUXILIARES ====================

def _get_fecha_sesion() -> date:
    raw = ctx_date.get()
    if raw:
        try:
            return date.fromisoformat(raw)
        except ValueError:
            pass
    return date.today()


def _parse_params(desde_str: str, hasta_str: str):
    try:
        fecini = date.fromisoformat(desde_str.strip()) if desde_str.strip() else None
    except ValueError:
        fecini = None
    try:
        fecfin = date.fromisoformat(hasta_str.strip()) if hasta_str.strip() else _get_fecha_sesion()
    except ValueError:
        fecfin = _get_fecha_sesion()
    return fecini, fecfin


def _build_subtitulo(fecini, fecfin: date) -> str:
    if fecini:
        return f"Del {fecini.strftime('%d/%m/%Y')} al {fecfin.strftime('%d/%m/%Y')}"
    return f"Hasta el {fecfin.strftime('%d/%m/%Y')}"


def _get_movimientos(empr: str, fecini, fecfin: date) -> list:
    """
    Devuelve lista de dicts con los movimientos de caja en el rango.
    Cada dict: {tipo, fecha, emprcodi, cta, nombre, concepto,
                salida, entrada, saldo}
    El primer elemento puede ser de tipo 'SA' (saldo anterior).

    Nota: saldo = SUM(cajahabe - cajadebe)
          Entradas (cajahabe) aumentan el saldo de caja.
          Salidas  (cajadebe) reducen el saldo de caja.
    """
    conn = sess_conns.get_conn(readonly=True)
    try:
        cur = conn.cursor(dictionary=True)

        # Saldo anterior (movimientos previos a fecini)
        if fecini:
            cur.execute(
                "SELECT COALESCE(SUM(cajahabe - cajadebe), 0) AS saldo_ant"
                "  FROM linacaja"
                " WHERE emprcodi = %s AND cajafech < %s",
                (empr, fecini),
            )
            saldo_ant = float(cur.fetchone()["saldo_ant"] or 0)
        else:
            saldo_ant = 0.0

        # Movimientos en el rango
        params = [empr]
        where  = ["c.emprcodi = %s"]
        if fecini:
            where.append("c.cajafech >= %s")
            params.append(fecini)
        where.append("c.cajafech <= %s")
        params.append(fecfin)

        sql = (
            "SELECT c.cajaid, c.emprcodi, c.cajafech,"
            "       c.cliecodi, c.provcodi,"
            "       c.cajaconc, c.cajadebe, c.cajahabe,"
            "       COALESCE(cli.cliename, '') AS cliename,"
            "       COALESCE(prov.provname, '') AS provname"
            "  FROM linacaja c"
            "  LEFT JOIN linaclie cli"
            "         ON cli.emprcodi = c.emprcodi AND cli.cliecodi = c.cliecodi AND c.cliecodi > 0"
            "  LEFT JOIN linaprov prov"
            "         ON prov.emprcodi = c.emprcodi AND prov.provcodi = c.provcodi AND c.provcodi > 0"
            " WHERE " + " AND ".join(where) +
            " ORDER BY c.cajafech ASC, c.cajaid ASC"
        )
        cur.execute(sql, params)
        rows_raw = cur.fetchall()
        cur.close()
    finally:
        sess_conns.release_conn(conn)

    movimientos = []
    saldo_running = saldo_ant

    # Fila de saldo anterior
    if abs(saldo_ant) > 0.005:
        fecha_sa = (fecini - timedelta(days=1)).strftime("%d/%m/%Y") if fecini else ""
        movimientos.append({
            "tipo":     "SA",
            "fecha":    fecha_sa,
            "emprcodi": "",
            "cta":      "",
            "nombre":   "SALDO ANTERIOR",
            "concepto": "",
            "salida":   0.0,
            "entrada":  0.0,
            "saldo":    saldo_ant,
        })

    for r in rows_raw:
        cajadebe = float(r["cajadebe"] or 0)
        cajahabe = float(r["cajahabe"] or 0)
        saldo_running += cajahabe - cajadebe

        cliecodi = int(r["cliecodi"] or 0)
        provcodi = int(r["provcodi"] or 0)
        if cliecodi > 0:
            cta    = f"{cliecodi:04d}"
            nombre = str(r["cliename"]).strip()
        elif provcodi > 0:
            cta    = f"{provcodi:04d}"
            nombre = str(r["provname"]).strip()
        else:
            cta    = ""
            nombre = ""

        cajafech  = r["cajafech"]
        fecha_str = cajafech.strftime("%d/%m/%Y") if cajafech else ""

        movimientos.append({
            "tipo":     "MV",
            "fecha":    fecha_str,
            "emprcodi": str(r["emprcodi"] or ""),
            "cta":      cta,
            "nombre":   nombre,
            "concepto": str(r["cajaconc"] or ""),
            "salida":   cajadebe,
            "entrada":  cajahabe,
            "saldo":    saldo_running,
        })

    return movimientos


def _fmt_z(val: float) -> str:
    return fmt_money(val) if abs(val) > 0.005 else ""


# ==================== RUTAS ====================

@router.get("/", response_class=HTMLResponse)
async def lina41_index(request: Request):
    Lina41.set_prog_code(PROG_CODE)
    user = Lina41.get_current_user(request)
    if not user:
        return RedirectResponse("/login")
    return Lina41.templates.TemplateResponse(
        "lina41/seleccion.html",
        {"request": request, "hoy": _get_fecha_sesion().isoformat()},
    )


@router.get("/pdf")
async def lina41_pdf(
    request: Request,
    desde:   str = Query(default=""),
    hasta:   str = Query(default=""),
):
    Lina41.set_prog_code(PROG_CODE)
    user = Lina41.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    empr_code, empr_name = Lina41.get_empr_info()
    empr = ctx_empr.get() or DEFAULT_EMPR_CODE
    fecini, fecfin = _parse_params(desde, hasta)

    movs = _get_movimientos(empr, fecini, fecfin)

    movs_pdf = [
        {
            "tipo":     m["tipo"],
            "fecha":    m["fecha"],
            "emprcodi": m["emprcodi"],
            "cta":      m["cta"],
            "nombre":   m["nombre"],
            "concepto": m["concepto"],
            "salida":   _fmt_z(m["salida"]),
            "entrada":  _fmt_z(m["entrada"]),
            "saldo":    fmt_money(m["saldo"]),
        }
        for m in movs
    ]

    template = pdf_jinja_env.get_template("lina41/main.html")
    html_str = template.render(
        app_name        = APP_CONFIG.get("app_name", ""),
        app_description = APP_CONFIG.get("app_description", ""),
        prog_code       = PROG_CODE,
        empr_code       = empr_code,
        empr_name       = empr_name,
        usuario         = user,
        titulo          = "Movimientos de Caja",
        subtitulo       = _build_subtitulo(fecini, fecfin),
        fecha           = date.today().strftime("%d/%m/%Y"),
        hora            = datetime.now().strftime("%H:%M"),
        movimientos     = movs_pdf,
    )

    pdf = HTML(string=html_str).write_pdf()
    return Response(
        content    = pdf,
        media_type = "application/pdf",
        headers    = {"Content-Disposition": "inline; filename=movimientos_caja.pdf"},
    )


@router.get("/xlsx")
async def lina41_xlsx(
    request: Request,
    desde:   str = Query(default=""),
    hasta:   str = Query(default=""),
):
    Lina41.set_prog_code(PROG_CODE)
    user = Lina41.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    empr_code, empr_name = Lina41.get_empr_info()
    empr = ctx_empr.get() or DEFAULT_EMPR_CODE
    fecini, fecfin = _parse_params(desde, hasta)
    subtitulo = _build_subtitulo(fecini, fecfin)

    movs = _get_movimientos(empr, fecini, fecfin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos Caja"

    ws.merge_cells("A1:H1")
    ws["A1"]      = "Movimientos de Caja"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:H2")
    ws["A2"]      = f"{APP_CONFIG.get('app_name', '')} — {empr_code} {empr_name}"
    ws["A2"].font = SUBTITLE_FONT

    ws.merge_cells("A3:H3")
    ws["A3"]      = (f"{subtitulo} — Usuario: {user} — "
                     f"{date.today().strftime('%d/%m/%Y')} {datetime.now().strftime('%H:%M')}")
    ws["A3"].font = SUBTITLE_FONT

    ws.append([])

    headers = ["Fecha", "Em.", "Cta.", "Nombre", "Concepto", "Salidas", "Entradas", "Saldo"]
    ws.append(headers)
    r = ws.max_row
    for cell in ws[r]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    SA_FILL = PatternFill(patternType="solid", fgColor="E9EFF9")
    SA_FONT = Font(name="Arial", size=8, italic=True)
    NUM_ALIGN = Alignment(horizontal="right")

    for m in movs:
        sal_val = m["salida"]  if abs(m["salida"])  > 0.005 else None
        ent_val = m["entrada"] if abs(m["entrada"]) > 0.005 else None
        ws.append([
            m["fecha"], m["emprcodi"], m["cta"], m["nombre"], m["concepto"],
            sal_val, ent_val, m["saldo"],
        ])
        r = ws.max_row
        for col_idx in (6, 7, 8):
            cell = ws.cell(r, col_idx)
            if cell.value is not None:
                cell.number_format = CURRENCY_FORMAT
            cell.alignment = NUM_ALIGN
        if m["tipo"] == "SA":
            for c in ws[r]:
                c.fill = SA_FILL
                c.font = SA_FONT

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 7
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(
        content    = buffer.read(),
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers    = {"Content-Disposition": "attachment; filename=movimientos_caja.xlsx"},
    )


@router.get("/txt")
async def lina41_txt(
    request: Request,
    desde:   str = Query(default=""),
    hasta:   str = Query(default=""),
):
    Lina41.set_prog_code(PROG_CODE)
    user = Lina41.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    empr_code, empr_name = Lina41.get_empr_info()
    empr = ctx_empr.get() or DEFAULT_EMPR_CODE
    fecini, fecfin = _parse_params(desde, hasta)
    subtitulo = _build_subtitulo(fecini, fecfin)

    movs = _get_movimientos(empr, fecini, fecfin)

    SEP = "=" * 88
    LINE = "-" * 88
    HDR = (f"{'Fecha':<10}  {'Em':2}  {'Cta':4}  {'Nombre':<20}  "
           f"{'Concepto':<20}  {'Salidas':>11}  {'Entradas':>11}  {'Saldo':>11}")

    lines = [
        "MOVIMIENTOS DE CAJA",
        subtitulo,
        f"{APP_CONFIG.get('app_name', '')} — {empr_code} {empr_name}",
        f"Usuario: {user} — {date.today().strftime('%d/%m/%Y')} {datetime.now().strftime('%H:%M')}",
        "",
        HDR,
        LINE,
    ]

    for m in movs:
        nombre_t  = m["nombre"][:20]
        concepto_t = m["concepto"][:20]
        if m["tipo"] == "SA":
            sal_s = "".rjust(11)
            ent_s = "".rjust(11)
            saldo_s = fmt_money(m["saldo"]).rjust(11)
            lines.append(
                f"{m['fecha']:<10}  {'':2}  {'':4}  {nombre_t:<20}  "
                f"{'':20}  {sal_s}  {ent_s}  {saldo_s}"
            )
        else:
            sal_s   = _fmt_z(m["salida"]).rjust(11)
            ent_s   = _fmt_z(m["entrada"]).rjust(11)
            saldo_s = fmt_money(m["saldo"]).rjust(11)
            lines.append(
                f"{m['fecha']:<10}  {m['emprcodi']:2}  {m['cta']:4}  {nombre_t:<20}  "
                f"{concepto_t:<20}  {sal_s}  {ent_s}  {saldo_s}"
            )

    lines.append(SEP)

    txt = "\r\n".join(lines)
    return Response(
        content    = txt.encode("utf-8"),
        media_type = "text/plain; charset=utf-8",
        headers    = {"Content-Disposition": "inline; filename=movimientos_caja.txt"},
    )
