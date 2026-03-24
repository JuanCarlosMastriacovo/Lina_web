from fastapi import APIRouter, Request, Query
from fastapi.responses import Response, HTMLResponse, RedirectResponse
from weasyprint import HTML
from jinja2 import Environment, FileSystemLoader
from pathlib import Path
from datetime import date, datetime
from io import BytesIO

import openpyxl

from CapaBRL.linabase import linabase
from CapaDAL.tablebase import get_table_model
from CapaDAL.config import APP_CONFIG
from CapaUI.xlsx_styles import (
    TITLE_FONT, SUBTITLE_FONT, HEADER_FONT, HEADER_FILL, HEADER_ALIGN,
)

# ==================== CONSTANTES Y ROUTER ====================

router     = APIRouter()
PROG_CODE  = "LINA122"
ROUTE_BASE = "/lina122"

LinaProv             = get_table_model("linaprov")
SUPPLIER_KEY_FIELD   = LinaProv.get_business_key_field()
SUPPLIER_LABEL_FIELD = LinaProv.get_selector_fields()[1]

pdf_templates_dir = Path(__file__).parent.parent / "templates"
pdf_jinja_env     = Environment(loader=FileSystemLoader(str(pdf_templates_dir)))


# ==================== CLASE PRINCIPAL ====================

class Lina122(linabase):
    """Módulo de listado de proveedores (LINA122)."""
    pass


# ==================== FUNCIONES AUXILIARES ====================

def _get_proveedores(conn, desde: int, hasta: int):
    """Obtiene proveedores filtrados por rango de código."""
    provs = LinaProv.list_all(
        order_by=SUPPLIER_KEY_FIELD,
        fields=[SUPPLIER_KEY_FIELD, SUPPLIER_LABEL_FIELD],
        conn=conn,
    )
    return [
        [str(p.get(SUPPLIER_KEY_FIELD) or "0").zfill(4), str(p.get(SUPPLIER_LABEL_FIELD) or "")]
        for p in provs
        if desde <= int(p.get(SUPPLIER_KEY_FIELD) or 0) <= hasta
    ]


# ==================== RUTAS ====================

@router.get("/", response_class=HTMLResponse)
async def lina122_index(request: Request):
    Lina122.set_prog_code(PROG_CODE)
    user = Lina122.get_current_user(request)
    if not user:
        return RedirectResponse("/login")
    return Lina122.templates.TemplateResponse(
        "lina122/seleccion.html",
        {"request": request, "error": None},
    )


@router.get("/pdf")
async def lina122_pdf(
    request: Request,
    desde: int = Query(default=0),
    hasta: int = Query(default=9999),
):
    """Genera el PDF del listado de proveedores."""
    Lina122.set_prog_code(PROG_CODE)
    user = Lina122.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    if desde > hasta:
        return Lina122.templates.TemplateResponse(
            "lina122/seleccion.html",
            {"request": request, "error": "El valor 'Desde' no puede ser mayor que 'Hasta'."},
        )

    empr_code, empr_name = Lina122.get_empr_info()
    conn  = Lina122.get_task_conn(request, readonly=True)
    filas = _get_proveedores(conn, desde, hasta)

    template = pdf_jinja_env.get_template("lina122/main.html")
    html_str = template.render(
        app_name        = APP_CONFIG.get("app_name", ""),
        app_description = APP_CONFIG.get("app_description", ""),
        prog_code = PROG_CODE,
        empr_code = empr_code,
        empr_name = empr_name,
        usuario   = user,
        titulo    = "Listado de Proveedores",
        subtitulo = f"Proveedores {str(desde).zfill(4)} al {str(hasta).zfill(4)}",
        fecha     = date.today().strftime("%d/%m/%Y"),
        hora      = datetime.now().strftime("%H:%M"),
        columnas  = ["Código", "Nombre"],
        filas     = filas,
    )

    pdf = HTML(string=html_str).write_pdf()
    return Response(
        content    = pdf,
        media_type = "application/pdf",
        headers    = {"Content-Disposition": "inline; filename=proveedores.pdf"},
    )


@router.get("/xlsx")
async def lina122_xlsx(
    request: Request,
    desde: int = Query(default=0),
    hasta: int = Query(default=9999),
):
    """Genera el XLSX del listado de proveedores."""
    Lina122.set_prog_code(PROG_CODE)
    user = Lina122.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    if desde > hasta:
        return Lina122.templates.TemplateResponse(
            "lina122/seleccion.html",
            {"request": request, "error": "El valor 'Desde' no puede ser mayor que 'Hasta'."},
        )

    empr_code, empr_name = Lina122.get_empr_info()
    conn  = Lina122.get_task_conn(request, readonly=True)
    filas = _get_proveedores(conn, desde, hasta)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proveedores"

    header_font   = HEADER_FONT
    header_fill   = HEADER_FILL
    header_align  = HEADER_ALIGN
    title_font    = TITLE_FONT
    subtitle_font = SUBTITLE_FONT

    ws.merge_cells("A1:B1")
    ws["A1"]      = "Listado de Proveedores"
    ws["A1"].font = title_font

    ws.merge_cells("A2:B2")
    ws["A2"]      = f"{APP_CONFIG.get('app_name', '')} — {empr_code} {empr_name}"
    ws["A2"].font = subtitle_font

    ws.merge_cells("A3:B3")
    ws["A3"]      = f"Proveedores {str(desde).zfill(4)} al {str(hasta).zfill(4)} — Usuario: {user} — {date.today().strftime('%d/%m/%Y')} {datetime.now().strftime('%H:%M')}"
    ws["A3"].font = subtitle_font

    ws.append([])

    ws.append(["Código", "Nombre"])
    for cell in ws[5]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    for fila in filas:
        ws.append(fila)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 45

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        content    = buffer.read(),
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers    = {"Content-Disposition": "attachment; filename=proveedores.xlsx"},
    )
