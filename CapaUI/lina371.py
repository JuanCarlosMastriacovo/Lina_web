from fastapi import APIRouter, Request, Query
from fastapi.responses import Response, HTMLResponse, RedirectResponse, JSONResponse
from weasyprint import HTML
from jinja2 import Environment, FileSystemLoader
from pathlib import Path
from datetime import date, datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

from CapaBRL.linabase import linabase
from CapaBRL.formatters import fmt_money
from CapaDAL.dataconn import sess_conns, ctx_empr, ctx_date
from CapaDAL.tablebase import get_table_model
from CapaBRL.config import APP_CONFIG
from CapaBRL.config import DEFAULT_EMPR_CODE
from CapaUI.xlsx_styles import (
    TITLE_FONT, SUBTITLE_FONT, HEADER_FONT, HEADER_FILL, HEADER_ALIGN, CURRENCY_FORMAT,
)

# ==================== CONSTANTES Y ROUTER ====================

router     = APIRouter()
PROG_CODE  = "LINA371"
ROUTE_BASE = "/lina371"

LinaProv = get_table_model("linaprov")

pdf_templates_dir = Path(__file__).parent.parent / "templates"
pdf_jinja_env     = Environment(loader=FileSystemLoader(str(pdf_templates_dir)))


# ==================== CLASE PRINCIPAL ====================

class Lina371(linabase):
    """Resúmenes de Cuenta Proveedores (LINA371)."""
    pass


# ==================== FUNCIONES AUXILIARES ====================

def _get_fecha_sesion() -> date:
    raw = ctx_date.get()
    if raw:
        try:
            return date.fromisoformat(raw)
        except ValueError:
            pass
    return date.today()


def _fmt_z(val: float) -> str:
    """Devuelve string vacío cuando el valor es cero."""
    return fmt_money(val) if abs(val) > 0.005 else ""


def _parse_params(desde_str: str, hasta_str: str, codiin_str: str, codifi_str: str):
    try:
        fecini = date.fromisoformat(desde_str.strip()) if desde_str.strip() else None
    except ValueError:
        fecini = None
    try:
        fecfin = date.fromisoformat(hasta_str.strip()) if hasta_str.strip() else _get_fecha_sesion()
    except ValueError:
        fecfin = _get_fecha_sesion()
    try:
        codiin = max(0, int(codiin_str))
    except (ValueError, TypeError):
        codiin = 0
    try:
        codifi = min(9999, int(codifi_str))
    except (ValueError, TypeError):
        codifi = 9999
    return fecini, fecfin, codiin, codifi


def _build_subtitulo(fecini, fecfin: date, codiin: int, codifi: int) -> str:
    s = f"Del {fecini.strftime('%d/%m/%Y')} al" if fecini else "Hasta el"
    s += f" {fecfin.strftime('%d/%m/%Y')} — Proveedores: {codiin:04d} a {codifi:04d}"
    return s


def _get_proveedores(empr: str, fecini, fecfin: date,
                     codiin: int, codifi: int, saldo_cero: bool) -> list:
    """
    Llama a sp_rescta_provs y devuelve lista de dicts por proveedor.
    Cada dict: {provcodi, provname, lineas: [{tipo, fecha, concepto, debe, habe, saldo}],
                saldo_final}

    Nota sobre debe/haber:
        ctprhabe = facturas recibidas  → se muestra en columna "Debe"  (aumenta lo que debemos)
        ctprdebe = pagos realizados    → se muestra en columna "Haber" (reduce lo que debemos)
        saldo_final = SUM(ctprhabe - ctprdebe) = positivo → debemos al proveedor
    """
    p_saldo_cero = 1 if saldo_cero else 0
    conn = sess_conns.get_conn(readonly=True)
    try:
        cur = conn.cursor(dictionary=True)
        cur.callproc("sp_rescta_provs",
                     [empr, fecini, fecfin, codiin, codifi, p_saldo_cero])
        rows = []
        for result in cur.stored_results():
            rows.extend(result.fetchall())
        cur.close()
    finally:
        sess_conns.release_conn(conn)

    proveedores = []
    current     = None

    for row in rows:
        provcodi = int(row["provcodi"])
        provname = str(row["provname"] or "").strip()

        if current is None or current["provcodi"] != provcodi:
            if current is not None:
                _save_proveedor(current, proveedores)
            current = {
                "provcodi":      provcodi,
                "provname":      provname,
                "saldo_running": 0.0,
                "lineas":        [],
                "saldo_final":   0.0,
            }

        if row["linea_tipo"] == "SA":
            saldo_ant = float(row["saldo_ant"] or 0)
            current["saldo_running"] = saldo_ant
            fech = row["ctprfech"]
            fecha_str = fech.strftime("%d/%m/%Y") if fech else ""
            if abs(saldo_ant) > 0.005:
                current["lineas"].append({
                    "tipo":     "SA",
                    "fecha":    fecha_str,
                    "concepto": "SALDO ANTERIOR",
                    "debe":     0.0,
                    "habe":     0.0,
                    "saldo":    saldo_ant,
                })
        else:
            # ctprhabe = factura (muestra en Debe); ctprdebe = pago (muestra en Haber)
            debe = float(row["ctprhabe"] or 0)
            habe = float(row["ctprdebe"] or 0)
            current["saldo_running"] += debe - habe
            fech = row["ctprfech"]
            fecha_str = fech.strftime("%d/%m/%Y") if fech else ""
            current["lineas"].append({
                "tipo":     "MV",
                "fecha":    fecha_str,
                "concepto": str(row["concepto"] or ""),
                "debe":     debe,
                "habe":     habe,
                "saldo":    current["saldo_running"],
            })

        current["saldo_final"] = current["saldo_running"]

    if current is not None:
        _save_proveedor(current, proveedores)

    return proveedores


def _save_proveedor(proveedor: dict, proveedores: list):
    if proveedor["lineas"]:
        proveedores.append(proveedor)


# ==================== RUTAS ====================

@router.get("/", response_class=HTMLResponse)
async def lina371_index(request: Request):
    Lina371.set_prog_code(PROG_CODE)
    user = Lina371.get_current_user(request)
    if not user:
        return RedirectResponse("/login")
    return Lina371.templates.TemplateResponse(
        "lina371/seleccion.html",
        {"request": request, "hoy": _get_fecha_sesion().isoformat()},
    )


@router.get("/prov/info")
async def lina371_prov_info(provcodi: str = Query(default="")):
    empr = ctx_empr.get() or DEFAULT_EMPR_CODE
    try:
        cod = int(provcodi)
    except (ValueError, TypeError):
        return JSONResponse({"ok": False, "provname": ""})
    rec = LinaProv.row_get({"emprcodi": empr, "provcodi": cod})
    if not rec:
        return JSONResponse({"ok": False, "provname": ""})
    return JSONResponse({"ok": True, "provname": str(rec.get("provname") or "").strip()})


@router.get("/pdf")
async def lina371_pdf(
    request:    Request,
    desde:      str = Query(default=""),
    hasta:      str = Query(default=""),
    codiin:     str = Query(default="0"),
    codifi:     str = Query(default="9999"),
    saldo_cero: int = Query(default=1),
):
    Lina371.set_prog_code(PROG_CODE)
    user = Lina371.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    empr_code, empr_name = Lina371.get_empr_info()
    empr = ctx_empr.get() or DEFAULT_EMPR_CODE
    fecini, fecfin, codiin_i, codifi_i = _parse_params(desde, hasta, codiin, codifi)

    proveedores = _get_proveedores(empr, fecini, fecfin, codiin_i, codifi_i, bool(saldo_cero))

    proveedores_pdf = [
        {
            "provcodi":   f"{p['provcodi']:04d}",
            "provname":   p["provname"],
            "saldo_final": fmt_money(p["saldo_final"]),
            "lineas": [
                {
                    "tipo":     ln["tipo"],
                    "fecha":    ln["fecha"],
                    "concepto": ln["concepto"],
                    "debe":     _fmt_z(ln["debe"]),
                    "habe":     _fmt_z(ln["habe"]),
                    "saldo":    fmt_money(ln["saldo"]),
                }
                for ln in p["lineas"]
            ],
        }
        for p in proveedores
    ]

    template = pdf_jinja_env.get_template("lina371/main.html")
    html_str  = template.render(
        app_name        = APP_CONFIG.get("app_name", ""),
        app_description = APP_CONFIG.get("app_description", ""),
        prog_code       = PROG_CODE,
        empr_code       = empr_code,
        empr_name       = empr_name,
        usuario         = user,
        titulo          = "Resumen de Cuenta Proveedores",
        subtitulo       = _build_subtitulo(fecini, fecfin, codiin_i, codifi_i),
        fecha           = date.today().strftime("%d/%m/%Y"),
        hora            = datetime.now().strftime("%H:%M"),
        proveedores     = proveedores_pdf,
    )

    pdf = HTML(string=html_str).write_pdf()
    return Response(
        content    = pdf,
        media_type = "application/pdf",
        headers    = {"Content-Disposition": "inline; filename=rescta_proveedores.pdf"},
    )


@router.get("/xlsx")
async def lina371_xlsx(
    request:    Request,
    desde:      str = Query(default=""),
    hasta:      str = Query(default=""),
    codiin:     str = Query(default="0"),
    codifi:     str = Query(default="9999"),
    saldo_cero: int = Query(default=1),
):
    Lina371.set_prog_code(PROG_CODE)
    user = Lina371.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    empr_code, empr_name = Lina371.get_empr_info()
    empr = ctx_empr.get() or DEFAULT_EMPR_CODE
    fecini, fecfin, codiin_i, codifi_i = _parse_params(desde, hasta, codiin, codifi)
    subtitulo = _build_subtitulo(fecini, fecfin, codiin_i, codifi_i)

    proveedores = _get_proveedores(empr, fecini, fecfin, codiin_i, codifi_i, bool(saldo_cero))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cta Proveedores"

    ws.merge_cells("A1:E1")
    ws["A1"]      = "Resumen de Cuenta Proveedores"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:E2")
    ws["A2"]      = f"{APP_CONFIG.get('app_name', '')} — {empr_code} {empr_name}"
    ws["A2"].font = SUBTITLE_FONT

    ws.merge_cells("A3:E3")
    ws["A3"]      = (f"{subtitulo} — Usuario: {user} — "
                     f"{date.today().strftime('%d/%m/%Y')} {datetime.now().strftime('%H:%M')}")
    ws["A3"].font = SUBTITLE_FONT

    PROV_FILL = PatternFill(patternType="solid", fgColor="4472C4")
    PROV_FONT = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    SA_FILL   = PatternFill(patternType="solid", fgColor="E9EFF9")
    SA_FONT   = Font(name="Arial", size=8, italic=True)
    NUM_ALIGN = Alignment(horizontal="right")

    for prov in proveedores:
        ws.append([])

        ws.append([f"{prov['provcodi']:04d}  {prov['provname']}"])
        r = ws.max_row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.cell(r, 1).fill = PROV_FILL
        ws.cell(r, 1).font = PROV_FONT

        # Columnas: Fecha | Concepto | Debe (factura) | Haber (pago) | Saldo
        ws.append(["Fecha", "Concepto", "Debe", "Haber", "Saldo"])
        r = ws.max_row
        for cell in ws[r]:
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = HEADER_ALIGN

        for ln in prov["lineas"]:
            debe_val = ln["debe"] if abs(ln["debe"]) > 0.005 else None
            habe_val = ln["habe"] if abs(ln["habe"]) > 0.005 else None
            ws.append([ln["fecha"], ln["concepto"], debe_val, habe_val, ln["saldo"]])
            r = ws.max_row
            for col_idx in (3, 4, 5):
                cell = ws.cell(r, col_idx)
                if cell.value is not None:
                    cell.number_format = CURRENCY_FORMAT
                cell.alignment = NUM_ALIGN
            if ln["tipo"] == "SA":
                for c in ws[r]:
                    c.fill = SA_FILL
                    c.font = SA_FONT

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(
        content    = buffer.read(),
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers    = {"Content-Disposition": "attachment; filename=rescta_proveedores.xlsx"},
    )


@router.get("/txt")
async def lina371_txt(
    request:    Request,
    desde:      str = Query(default=""),
    hasta:      str = Query(default=""),
    codiin:     str = Query(default="0"),
    codifi:     str = Query(default="9999"),
    saldo_cero: int = Query(default=1),
):
    Lina371.set_prog_code(PROG_CODE)
    user = Lina371.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    empr_code, empr_name = Lina371.get_empr_info()
    empr = ctx_empr.get() or DEFAULT_EMPR_CODE
    fecini, fecfin, codiin_i, codifi_i = _parse_params(desde, hasta, codiin, codifi)
    subtitulo = _build_subtitulo(fecini, fecfin, codiin_i, codifi_i)

    proveedores = _get_proveedores(empr, fecini, fecfin, codiin_i, codifi_i, bool(saldo_cero))

    SEP  = "=" * 76
    LINE = "-" * 76
    HDR  = f"{'Fecha':<10}  {'Concepto':<22}  {'Debe':>12}  {'Haber':>12}  {'Saldo':>12}"

    lines = [
        "RESUMEN DE CUENTA PROVEEDORES",
        subtitulo,
        f"{APP_CONFIG.get('app_name', '')} — {empr_code} {empr_name}",
        f"Usuario: {user} — {date.today().strftime('%d/%m/%Y')} {datetime.now().strftime('%H:%M')}",
        "",
    ]

    for prov in proveedores:
        encab = f"{prov['provcodi']:04d}  {prov['provname']}"
        pad   = (76 - len(encab)) // 2
        lines.append(" " * max(pad, 0) + encab)
        lines.append(LINE)
        lines.append(HDR)
        lines.append(LINE)
        for ln in prov["lineas"]:
            debe_s  = _fmt_z(ln["debe"]).rjust(12)
            habe_s  = _fmt_z(ln["habe"]).rjust(12)
            saldo_s = fmt_money(ln["saldo"]).rjust(12)
            lines.append(
                f"{ln['fecha']:<10}  {ln['concepto']:<22}  {debe_s}  {habe_s}  {saldo_s}"
            )
        lines.append(SEP)
        lines.append("")

    txt = "\r\n".join(lines)
    return Response(
        content    = txt.encode("utf-8"),
        media_type = "text/plain; charset=utf-8",
        headers    = {"Content-Disposition": "inline; filename=rescta_proveedores.txt"},
    )
