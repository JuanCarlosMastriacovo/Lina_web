from fastapi import APIRouter, Request, Query
from fastapi.responses import Response, HTMLResponse, RedirectResponse, JSONResponse
from weasyprint import HTML
from jinja2 import Environment, FileSystemLoader
from pathlib import Path
from datetime import date, datetime
from io import BytesIO

import openpyxl

from CapaBRL.linabase import linabase
from CapaBRL.formatters import fmt_money
from CapaBRL.txt_brl import generar_txt, col as txt_col
from CapaDAL.dataconn import sess_conns, ctx_empr
from CapaDAL.tablebase import get_table_model
from CapaBRL.config import APP_CONFIG
from CapaBRL.config import DEFAULT_EMPR_CODE
from CapaUI.xlsx_styles import (
    TITLE_FONT, SUBTITLE_FONT, HEADER_FONT, HEADER_FILL, HEADER_ALIGN,
    TOTAL_FONT, CURRENCY_FORMAT,
)

# ==================== CONSTANTES Y ROUTER ====================

router     = APIRouter()
PROG_CODE  = "LINA231"
ROUTE_BASE = "/lina231"

LinaClie = get_table_model("linaclie")

pdf_templates_dir = Path(__file__).parent.parent / "templates"
pdf_jinja_env     = Environment(loader=FileSystemLoader(str(pdf_templates_dir)))


# ==================== CLASE PRINCIPAL ====================

class Lina231(linabase):
    """Listado Resumen de Ventas (LINA231)."""
    pass


# ==================== FUNCIONES AUXILIARES ====================

def _parse_fechas(desde_str: str, hasta_str: str):
    try:
        fecha_desde = date.fromisoformat(desde_str.strip()) if desde_str.strip() else None
    except ValueError:
        fecha_desde = None
    try:
        fecha_hasta = date.fromisoformat(hasta_str.strip()) if hasta_str.strip() else date.today()
    except ValueError:
        fecha_hasta = date.today()
    return fecha_desde, fecha_hasta


def _build_subtitulo(fecha_desde, fecha_hasta: date, cliecodi: int) -> str:
    if fecha_desde:
        s = f"Del {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')}"
    else:
        s = f"Hasta el {fecha_hasta.strftime('%d/%m/%Y')}"
    if cliecodi > 0:
        s += f" — Cliente: {cliecodi:04d}"
    return s


def _get_filas(empr: str, fecha_desde, fecha_hasta: date, cliecodi: int) -> list:
    """
    Retorna lista de [codm, nro, fecha_str, cta_str, cliente, importe_float].
    Ordenado por fecha ASC, nro ASC (igual que el reporte FoxPro).
    """
    conn = sess_conns.get_conn(readonly=True)
    try:
        cur = conn.cursor(dictionary=True)

        params = [empr]
        where  = ["fvhe.emprcodi = %s"]

        if fecha_desde:
            where.append("fvhe.fvhefech >= %s")
            params.append(fecha_desde)
        where.append("fvhe.fvhefech <= %s")
        params.append(fecha_hasta)

        if cliecodi > 0:
            where.append("fvhe.cliecodi = %s")
            params.append(cliecodi)

        sql = (
            "SELECT fvhe.codmcodi, fvhe.fvhenume, fvhe.fvhefech,"
            "       fvhe.cliecodi, fvhe.fvhetota, fvhe.fvheobse,"
            "       COALESCE(clie.cliename, '') AS cliename"
            "  FROM linafvhe fvhe"
            "  LEFT JOIN linaclie clie"
            "         ON clie.emprcodi = fvhe.emprcodi"
            "        AND clie.cliecodi = fvhe.cliecodi"
            " WHERE " + " AND ".join(where) +
            " ORDER BY fvhe.fvhefech ASC, fvhe.fvhenume ASC"
        )
        cur.execute(sql, params)
        rows_raw = cur.fetchall()
        cur.close()
    finally:
        sess_conns.release_conn(conn)

    filas = []
    for r in rows_raw:
        obse    = str(r.get("fvheobse") or "").strip()
        anulado = obse.startswith("*** ANULAD")
        cliente = obse if anulado else str(r.get("cliename") or "").strip()
        fvhefech = r.get("fvhefech")
        fecha_str = fvhefech.strftime("%d/%m/%Y") if hasattr(fvhefech, "strftime") else str(fvhefech or "")
        filas.append([
            str(r.get("codmcodi") or ""),
            int(r.get("fvhenume") or 0),
            fecha_str,
            str(int(r.get("cliecodi") or 0)).zfill(4),
            cliente,
            float(r.get("fvhetota") or 0),
        ])
    return filas


# ==================== RUTAS ====================

@router.get("/", response_class=HTMLResponse)
async def lina231_index(request: Request):
    Lina231.set_prog_code(PROG_CODE)
    user = Lina231.get_current_user(request)
    if not user:
        return RedirectResponse("/login")
    return Lina231.templates.TemplateResponse(
        "lina231/seleccion.html",
        {"request": request, "hoy": date.today().isoformat()},
    )


@router.get("/clie/info")
async def lina231_clie_info(cliecodi: str = Query(default="")):
    empr = ctx_empr.get() or DEFAULT_EMPR_CODE
    try:
        cod = int(cliecodi)
    except (ValueError, TypeError):
        return JSONResponse({"ok": False, "cliename": ""})
    rec = LinaClie.row_get({"emprcodi": empr, "cliecodi": cod})
    if not rec:
        return JSONResponse({"ok": False, "cliename": ""})
    return JSONResponse({"ok": True, "cliename": str(rec.get("cliename") or "").strip()})


@router.get("/pdf")
async def lina231_pdf(
    request:  Request,
    desde:    str = Query(default=""),
    hasta:    str = Query(default=""),
    cliecodi: int = Query(default=0),
):
    Lina231.set_prog_code(PROG_CODE)
    user = Lina231.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    empr_code, empr_name = Lina231.get_empr_info()
    empr = ctx_empr.get() or DEFAULT_EMPR_CODE
    fecha_desde, fecha_hasta = _parse_fechas(desde, hasta)

    filas      = _get_filas(empr, fecha_desde, fecha_hasta, cliecodi)
    total_impo = sum(f[5] for f in filas)

    filas_pdf = [
        [f[0], str(f[1]).zfill(6), f[2], f[3], f[4], fmt_money(f[5])]
        for f in filas
    ]

    template = pdf_jinja_env.get_template("lina231/main.html")
    html_str = template.render(
        app_name        = APP_CONFIG.get("app_name", ""),
        app_description = APP_CONFIG.get("app_description", ""),
        prog_code       = PROG_CODE,
        empr_code       = empr_code,
        empr_name       = empr_name,
        usuario         = user,
        titulo          = "Listado Resumen de Ventas",
        subtitulo       = _build_subtitulo(fecha_desde, fecha_hasta, cliecodi),
        fecha           = date.today().strftime("%d/%m/%Y"),
        hora            = datetime.now().strftime("%H:%M"),
        filas           = filas_pdf,
        total_importe   = fmt_money(total_impo),
    )

    pdf = HTML(string=html_str).write_pdf()
    return Response(
        content    = pdf,
        media_type = "application/pdf",
        headers    = {"Content-Disposition": "inline; filename=resumen_ventas.pdf"},
    )


@router.get("/xlsx")
async def lina231_xlsx(
    request:  Request,
    desde:    str = Query(default=""),
    hasta:    str = Query(default=""),
    cliecodi: int = Query(default=0),
):
    Lina231.set_prog_code(PROG_CODE)
    user = Lina231.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    empr_code, empr_name = Lina231.get_empr_info()
    empr = ctx_empr.get() or DEFAULT_EMPR_CODE
    fecha_desde, fecha_hasta = _parse_fechas(desde, hasta)

    filas      = _get_filas(empr, fecha_desde, fecha_hasta, cliecodi)
    total_impo = sum(f[5] for f in filas)
    subtitulo  = _build_subtitulo(fecha_desde, fecha_hasta, cliecodi)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen Ventas"

    ws.merge_cells("A1:F1")
    ws["A1"]      = "Listado Resumen de Ventas"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:F2")
    ws["A2"]      = f"{APP_CONFIG.get('app_name', '')} — {empr_code} {empr_name}"
    ws["A2"].font = SUBTITLE_FONT

    ws.merge_cells("A3:F3")
    ws["A3"]      = (f"{subtitulo} — Usuario: {user} — "
                     f"{date.today().strftime('%d/%m/%Y')} {datetime.now().strftime('%H:%M')}")
    ws["A3"].font = SUBTITLE_FONT

    ws.append([])  # fila 4 vacía

    headers = ["Cpbt", "Número", "Fecha", "Cta.", "Cliente", "Importe"]
    ws.append(headers)
    for cell in ws[5]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    DATA_START = 6
    for fila in filas:
        ws.append([
            fila[0],
            str(fila[1]).zfill(6),
            fila[2],
            fila[3],
            fila[4],
            fila[5],
        ])
        ws.cell(row=ws.max_row, column=6).number_format = CURRENCY_FORMAT

    total_row = DATA_START + len(filas)
    ws.append(["", "", "", "", "TOTAL", total_impo])
    for cell in ws[total_row]:
        cell.font = TOTAL_FONT
    ws.cell(row=total_row, column=6).number_format = CURRENCY_FORMAT

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 6
    ws.column_dimensions["E"].width = 42
    ws.column_dimensions["F"].width = 14

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        content    = buffer.read(),
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers    = {"Content-Disposition": "attachment; filename=resumen_ventas.xlsx"},
    )


_COLUMNAS_TXT = [
    txt_col("Cpbt",    5,  "L"),
    txt_col("Número",  8,  "R"),
    txt_col("Fecha",   10, "C"),
    txt_col("Cta.",    5,  "R"),
    txt_col("Cliente", 38, "L"),
    txt_col("Importe", 15, "R"),
]


@router.get("/txt")
async def lina231_txt(
    request:  Request,
    desde:    str = Query(default=""),
    hasta:    str = Query(default=""),
    cliecodi: int = Query(default=0),
):
    Lina231.set_prog_code(PROG_CODE)
    user = Lina231.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    empr_code, empr_name = Lina231.get_empr_info()
    empr = ctx_empr.get() or DEFAULT_EMPR_CODE
    fecha_desde, fecha_hasta = _parse_fechas(desde, hasta)

    filas      = _get_filas(empr, fecha_desde, fecha_hasta, cliecodi)
    total_impo = sum(f[5] for f in filas)

    filas_txt = [
        [f[0], str(f[1]).zfill(6), f[2], f[3], f[4], fmt_money(f[5])]
        for f in filas
    ]
    totales_txt = ["", "", "", "", "TOTAL", fmt_money(total_impo)]

    txt = generar_txt(
        titulo    = "Listado Resumen de Ventas",
        subtitulo = _build_subtitulo(fecha_desde, fecha_hasta, cliecodi),
        columnas  = _COLUMNAS_TXT,
        filas     = filas_txt,
        app_name  = APP_CONFIG.get("app_name", ""),
        empr_code = empr_code,
        empr_name = empr_name,
        usuario   = user,
        fecha     = date.today().strftime("%d/%m/%Y"),
        hora      = datetime.now().strftime("%H:%M"),
        totales   = totales_txt,
    )

    return Response(
        content    = txt.encode("utf-8"),
        media_type = "text/plain; charset=utf-8",
        headers    = {"Content-Disposition": "inline; filename=resumen_ventas.txt"},
    )
