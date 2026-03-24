from openpyxl.styles import Font, PatternFill, Alignment

# ==================== FUENTES ====================

TITLE_FONT    = Font(bold=True, size=12)
SUBTITLE_FONT = Font(size=8, color="555555")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=9)
TOTAL_FONT    = Font(bold=True, size=9)
GROUP_FONT    = Font(bold=True, size=9, color="1F3864")   # encabezado de grupo (lina1332)

# ==================== RELLENOS ====================

HEADER_FILL   = PatternFill("solid", fgColor="4472C4")
GROUP_FILL    = PatternFill("solid", fgColor="BDD7EE")    # grupo (lina1332)

# ==================== ALINEACIONES ====================

HEADER_ALIGN  = Alignment(horizontal="left", vertical="center")
RIGHT_ALIGN   = Alignment(horizontal="right")

# ==================== FORMATOS DE CELDA ====================

CURRENCY_FORMAT = '#,##0.00;-#,##0.00'
DATE_FMT  = 'DD/MM/YYYY'
