from fastapi import APIRouter, Request, Query
from fastapi.responses import Response, HTMLResponse, RedirectResponse, JSONResponse
from weasyprint import HTML
from jinja2 import Environment, FileSystemLoader
from pathlib import Path
from datetime import date, datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment

from CapaBRL.linabase import linabase
from CapaBRL.formatters import fmt_money
from CapaDAL.dataconn import sess_conns, ctx_empr
from CapaDAL.tablebase import get_table_model
from CapaDAL.config import APP_CONFIG
from CapaBRL.config import DEFAULT_EMPR_CODE
from CapaUI.xlsx_styles import (
    TITLE_FONT, SUBTITLE_FONT, HEADER_FONT, HEADER_FILL, HEADER_ALIGN,
    TOTAL_FONT, CURRENCY_FORMAT,
)

# ==================== CONSTANTES Y ROUTER ====================

router     = APIRouter()
PROG_CODE  = "LINA272"
ROUTE_BASE = "/lina272"

LinaClie = get_table_model("linaclie")

pdf_templates_dir = Path(__file__).parent.parent / "templates"
pdf_jinja_env     = Environment(loader=FileSystemLoader(str(pdf_templates_dir)))


# ==================== CLASE PRINCIPAL ====================

class Lina272(linabase):
    """Listado de Saldos de Clientes (LINA272)."""
    pass


# ==================== FUNCIONES AUXILIARES ====================

def _get_fecha_sesion() -> date:
    from CapaDAL.dataconn import ctx_date
    raw = ctx_date.get()
    if raw:
        try:
            return date.fromisoformat(raw)
        except ValueError:
            pass
    return date.today()


def _parse_params(hasta_str: str, codiin_str: str, codifi_str: str):
    try:
        fecfin = date.fromisoformat(hasta_str.strip()) if hasta_str.strip() else _get_fecha_sesion()
    except ValueError:
        fecfin = _get_fecha_sesion()
    try:
        codiin = max(0, int(codiin_str))
    except (ValueError, TypeError):
        codiin = 0
    try:
        codifi = min(9999, int(codifi_str))
    except (ValueError, TypeError):
        codifi = 9999
    return fecfin, codiin, codifi


def _build_subtitulo(fecfin: date, codiin: int, codifi: int) -> str:
    return (f"Saldos al {fecfin.strftime('%d/%m/%Y')} — "
            f"Clientes: {codiin:04d} a {codifi:04d}")


def _get_filas(empr: str, fecfin: date, codiin: int, codifi: int) -> list:
    """
    Llama a sp_saldo_clies y devuelve lista de dicts {cliecodi, cliename, saldo}.
    """
    conn = sess_conns.get_conn(readonly=True)
    try:
        cur = conn.cursor(dictionary=True)
        cur.callproc("sp_saldo_clies", [empr, codiin, codifi, fecfin])
        rows = []
        for result in cur.stored_results():
            rows.extend(result.fetchall())
        cur.close()
    finally:
        sess_conns.release_conn(conn)

    return [
        {
            "cliecodi": int(row["cliecodi"]),
            "cliename": str(row["cliename"] or "").strip(),
            "saldo":    float(row["saldo"] or 0),
        }
        for row in rows
    ]


# ==================== RUTAS ====================

@router.get("/", response_class=HTMLResponse)
async def lina272_index(request: Request):
    Lina272.set_prog_code(PROG_CODE)
    user = Lina272.get_current_user(request)
    if not user:
        return RedirectResponse("/login")
    return Lina272.templates.TemplateResponse(
        "lina272/seleccion.html",
        {"request": request, "hoy": _get_fecha_sesion().isoformat()},
    )


@router.get("/clie/info")
async def lina272_clie_info(cliecodi: str = Query(default="")):
    empr = ctx_empr.get() or DEFAULT_EMPR_CODE
    try:
        cod = int(cliecodi)
    except (ValueError, TypeError):
        return JSONResponse({"ok": False, "cliename": ""})
    rec = LinaClie.row_get({"emprcodi": empr, "cliecodi": cod})
    if not rec:
        return JSONResponse({"ok": False, "cliename": ""})
    return JSONResponse({"ok": True, "cliename": str(rec.get("cliename") or "").strip()})


@router.get("/pdf")
async def lina272_pdf(
    request: Request,
    hasta:   str = Query(default=""),
    codiin:  str = Query(default="0"),
    codifi:  str = Query(default="9999"),
):
    Lina272.set_prog_code(PROG_CODE)
    user = Lina272.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    empr_code, empr_name = Lina272.get_empr_info()
    empr = ctx_empr.get() or DEFAULT_EMPR_CODE
    fecfin, codiin_i, codifi_i = _parse_params(hasta, codiin, codifi)

    filas = _get_filas(empr, fecfin, codiin_i, codifi_i)
    total = sum(f["saldo"] for f in filas)

    filas_pdf = [
        {
            "cliecodi": f"{f['cliecodi']:04d}",
            "cliename": f["cliename"],
            "saldo":    fmt_money(f["saldo"]),
        }
        for f in filas
    ]

    template = pdf_jinja_env.get_template("lina272/main.html")
    html_str = template.render(
        app_name        = APP_CONFIG.get("app_name", ""),
        app_description = APP_CONFIG.get("app_description", ""),
        prog_code       = PROG_CODE,
        empr_code       = empr_code,
        empr_name       = empr_name,
        usuario         = user,
        titulo          = "Listado de Saldos de Clientes",
        subtitulo       = _build_subtitulo(fecfin, codiin_i, codifi_i),
        fecha           = date.today().strftime("%d/%m/%Y"),
        hora            = datetime.now().strftime("%H:%M"),
        filas           = filas_pdf,
        total           = fmt_money(total),
        cant            = len(filas),
    )

    pdf = HTML(string=html_str).write_pdf()
    return Response(
        content    = pdf,
        media_type = "application/pdf",
        headers    = {"Content-Disposition": "inline; filename=saldo_clientes.pdf"},
    )


@router.get("/xlsx")
async def lina272_xlsx(
    request: Request,
    hasta:   str = Query(default=""),
    codiin:  str = Query(default="0"),
    codifi:  str = Query(default="9999"),
):
    Lina272.set_prog_code(PROG_CODE)
    user = Lina272.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    empr_code, empr_name = Lina272.get_empr_info()
    empr = ctx_empr.get() or DEFAULT_EMPR_CODE
    fecfin, codiin_i, codifi_i = _parse_params(hasta, codiin, codifi)
    subtitulo = _build_subtitulo(fecfin, codiin_i, codifi_i)

    filas = _get_filas(empr, fecfin, codiin_i, codifi_i)
    total = sum(f["saldo"] for f in filas)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Saldos Clientes"

    # Bloque de título
    ws.merge_cells("A1:C1")
    ws["A1"]      = "Listado de Saldos de Clientes"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:C2")
    ws["A2"]      = f"{APP_CONFIG.get('app_name', '')} — {empr_code} {empr_name}"
    ws["A2"].font = SUBTITLE_FONT

    ws.merge_cells("A3:C3")
    ws["A3"]      = (f"{subtitulo} — Usuario: {user} — "
                     f"{date.today().strftime('%d/%m/%Y')} {datetime.now().strftime('%H:%M')}")
    ws["A3"].font = SUBTITLE_FONT

    ws.append([])

    # Cabecera de columnas
    ws.append(["Código", "Cliente", "Saldo"])
    r = ws.max_row
    for cell in ws[r]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    ws.cell(r, 3).alignment = Alignment(horizontal="right")

    NUM_ALIGN = Alignment(horizontal="right")

    for f in filas:
        ws.append([f["cliecodi"], f["cliename"], f["saldo"]])
        r = ws.max_row
        ws.cell(r, 1).number_format = "0000"
        ws.cell(r, 3).number_format = CURRENCY_FORMAT
        ws.cell(r, 3).alignment     = NUM_ALIGN

    # Fila de totales
    ws.append([f"{len(filas)} clientes", "", total])
    r = ws.max_row
    for cell in ws[r]:
        cell.font = TOTAL_FONT
    ws.cell(r, 3).number_format = CURRENCY_FORMAT
    ws.cell(r, 3).alignment     = NUM_ALIGN

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 16

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(
        content    = buffer.read(),
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers    = {"Content-Disposition": "attachment; filename=saldo_clientes.xlsx"},
    )


@router.get("/txt")
async def lina272_txt(
    request: Request,
    hasta:   str = Query(default=""),
    codiin:  str = Query(default="0"),
    codifi:  str = Query(default="9999"),
):
    Lina272.set_prog_code(PROG_CODE)
    user = Lina272.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    empr_code, empr_name = Lina272.get_empr_info()
    empr = ctx_empr.get() or DEFAULT_EMPR_CODE
    fecfin, codiin_i, codifi_i = _parse_params(hasta, codiin, codifi)
    subtitulo = _build_subtitulo(fecfin, codiin_i, codifi_i)

    filas = _get_filas(empr, fecfin, codiin_i, codifi_i)
    total = sum(f["saldo"] for f in filas)

    SEP  = "-" * 60
    HDR  = f"{'Cód':>4}  {'Cliente':<40}  {'Saldo':>12}"

    lines = [
        "LISTADO DE SALDOS DE CLIENTES",
        subtitulo,
        f"{APP_CONFIG.get('app_name', '')} — {empr_code} {empr_name}",
        f"Usuario: {user} — {date.today().strftime('%d/%m/%Y')} {datetime.now().strftime('%H:%M')}",
        "",
        HDR,
        SEP,
    ]

    for f in filas:
        lines.append(
            f"{f['cliecodi']:04d}  {f['cliename']:<40}  {fmt_money(f['saldo']):>12}"
        )

    lines += [
        SEP,
        f"{'Total':>46}  {fmt_money(total):>12}",
        f"{len(filas)} clientes",
    ]

    txt = "\r\n".join(lines)
    return Response(
        content    = txt.encode("utf-8"),
        media_type = "text/plain; charset=utf-8",
        headers    = {"Content-Disposition": "inline; filename=saldo_clientes.txt"},
    )
