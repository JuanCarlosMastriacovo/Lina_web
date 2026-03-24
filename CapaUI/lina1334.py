from fastapi import APIRouter, Request, Query
from fastapi.responses import Response, HTMLResponse, RedirectResponse
from weasyprint import HTML
from jinja2 import Environment, FileSystemLoader
from pathlib import Path
from datetime import date, datetime
from io import BytesIO

import openpyxl

from CapaBRL.linabase import linabase
from CapaBRL.formatters import fmt_money
from CapaBRL.stock_brl import get_existencias_batch
from CapaBRL.txt_brl import generar_txt, col as txt_col
from CapaDAL.tablebase import get_table_model
from CapaDAL.config import APP_CONFIG
from CapaUI.xlsx_styles import (
    TITLE_FONT, SUBTITLE_FONT, HEADER_FONT, HEADER_FILL, HEADER_ALIGN,
    TOTAL_FONT, CURRENCY_FORMAT,
)

# ==================== CONSTANTES Y ROUTER ====================

router     = APIRouter()
PROG_CODE  = "LINA1334"
ROUTE_BASE = "/lina1334"

LinaArti          = get_table_model("linaarti")
ARTICLE_KEY_FIELD = LinaArti.get_business_key_field()

pdf_templates_dir = Path(__file__).parent.parent / "templates"
pdf_jinja_env     = Environment(loader=FileSystemLoader(str(pdf_templates_dir)))


# ==================== CLASE PRINCIPAL ====================

class Lina1334(linabase):
    """Módulo de artículos bajo PMP (LINA1334)."""
    pass


# ==================== FUNCIONES AUXILIARES ====================

def _get_filas(fecha_hasta: date) -> list:
    """
    Retorna filas [artrcodi, articodi, artidesc, exis, artipmpe, faltan, artiprec, valor]
    solo para artículos con artipmpe > 0 y exis < artipmpe, ordenados por artrcodi+articodi.
    """
    arts = LinaArti.list_all(
        order_by=ARTICLE_KEY_FIELD,
        fields=["articodi", "artidesc", "artrcodi", "artiprec", "artipmpe"],
    )
    if not arts:
        return []

    articodis = [str(a.get("articodi") or "").strip() for a in arts]
    exist_map = get_existencias_batch(articodis, fecha_hasta)

    arts_sorted = sorted(arts, key=lambda a: (
        str(a.get("artrcodi") or "").strip(),
        str(a.get("articodi") or "").strip(),
    ))

    filas = []
    for a in arts_sorted:
        articodi = str(a.get("articodi") or "").strip()
        artrcodi = str(a.get("artrcodi") or "").strip()
        artidesc = str(a.get("artidesc") or "")
        artiprec = float(a.get("artiprec") or 0)
        artipmpe = int(a.get("artipmpe") or 0)
        exis     = exist_map.get(articodi, 0)

        if artipmpe <= 0 or exis >= artipmpe:
            continue

        faltan = artipmpe - exis
        valor  = faltan * artiprec
        filas.append([artrcodi, articodi, artidesc, exis, artipmpe, faltan, artiprec, valor])

    return filas


# ==================== RUTAS ====================

@router.get("/", response_class=HTMLResponse)
async def lina1334_index(request: Request):
    Lina1334.set_prog_code(PROG_CODE)
    user = Lina1334.get_current_user(request)
    if not user:
        return RedirectResponse("/login")
    return Lina1334.templates.TemplateResponse(
        "lina1334/seleccion.html",
        {"request": request, "error": None, "hoy": date.today().isoformat()},
    )


@router.get("/pdf")
async def lina1334_pdf(
    request: Request,
    fecha_hasta: str = Query(default=""),
):
    Lina1334.set_prog_code(PROG_CODE)
    user = Lina1334.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    try:
        fh = date.fromisoformat(fecha_hasta) if fecha_hasta else date.today()
    except ValueError:
        fh = date.today()

    empr_code, empr_name = Lina1334.get_empr_info()
    filas = _get_filas(fh)

    total_valor = sum(f[7] for f in filas)

    filas_pdf = [
        [f[0], f[1], f[2], f[3], f[4], f[5], fmt_money(f[6]), fmt_money(f[7])]
        for f in filas
    ]

    template = pdf_jinja_env.get_template("lina1334/main.html")
    html_str = template.render(
        app_name        = APP_CONFIG.get("app_name", ""),
        app_description = APP_CONFIG.get("app_description", ""),
        prog_code    = PROG_CODE,
        empr_code    = empr_code,
        empr_name    = empr_name,
        usuario      = user,
        titulo       = "Artículos Bajo Punto Mínimo de Pedido",
        fecha_hasta  = fh.strftime("%d/%m/%Y"),
        fecha        = date.today().strftime("%d/%m/%Y"),
        hora         = datetime.now().strftime("%H:%M"),
        filas        = filas_pdf,
        total_valor  = fmt_money(total_valor),
    )

    pdf = HTML(string=html_str).write_pdf()
    return Response(
        content    = pdf,
        media_type = "application/pdf",
        headers    = {"Content-Disposition": "inline; filename=articulos_bajo_pmp.pdf"},
    )


@router.get("/xlsx")
async def lina1334_xlsx(
    request: Request,
    fecha_hasta: str = Query(default=""),
):
    Lina1334.set_prog_code(PROG_CODE)
    user = Lina1334.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    try:
        fh = date.fromisoformat(fecha_hasta) if fecha_hasta else date.today()
    except ValueError:
        fh = date.today()

    empr_code, empr_name = Lina1334.get_empr_info()
    filas = _get_filas(fh)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bajo PMP"

    title_font    = TITLE_FONT
    subtitle_font = SUBTITLE_FONT
    header_font   = HEADER_FONT
    header_fill   = HEADER_FILL
    header_align  = HEADER_ALIGN
    total_font    = TOTAL_FONT

    ncols    = 8
    last_col = openpyxl.utils.get_column_letter(ncols)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"]      = "Artículos Bajo Punto Mínimo de Pedido"
    ws["A1"].font = title_font

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"]      = f"{APP_CONFIG.get('app_name', '')} — {empr_code} {empr_name}"
    ws["A2"].font = subtitle_font

    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"]      = (f"Al {fh.strftime('%d/%m/%Y')} — "
                     f"Usuario: {user} — "
                     f"{date.today().strftime('%d/%m/%Y')} {datetime.now().strftime('%H:%M')}")
    ws["A3"].font = subtitle_font

    ws.append([])  # fila 4 vacía

    headers = ["Rubro", "Código", "Descripción", "Existencia", "PMP", "Faltan", "Precio", "Valor"]
    ws.append(headers)   # fila 5
    for cell in ws[5]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    DATA_START = 6
    for i, fila in enumerate(filas):
        row_num = DATA_START + i
        # A=artrcodi B=articodi C=artidesc D=exis E=artipmpe F=faltan G=artiprec H=valor
        ws.append([fila[0], fila[1], fila[2], fila[3], fila[4], None, fila[6], None])
        ws.cell(row=row_num, column=6).value         = f"=E{row_num}-D{row_num}"
        ws.cell(row=row_num, column=7).number_format = CURRENCY_FORMAT
        ws.cell(row=row_num, column=8).value         = f"=F{row_num}*G{row_num}"
        ws.cell(row=row_num, column=8).number_format = CURRENCY_FORMAT

    total_row = DATA_START + len(filas)
    last_data = total_row - 1
    ws.append(["", "", "TOTAL", "", "", "", "", f"=SUM(H{DATA_START}:H{last_data})"])
    for cell in ws[total_row]:
        cell.font = total_font
    ws.cell(row=total_row, column=8).number_format = CURRENCY_FORMAT

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        content    = buffer.read(),
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers    = {"Content-Disposition": "attachment; filename=articulos_bajo_pmp.xlsx"},
    )


_COLUMNAS_TXT = [
    txt_col("Rubro",      6,  "L"),
    txt_col("Código",     9,  "L"),
    txt_col("Descripción",32, "L"),
    txt_col("Exist.",     8,  "R"),
    txt_col("PMP",        6,  "R"),
    txt_col("Faltan",     7,  "R"),
    txt_col("Precio",     12, "R"),
    txt_col("Valor",      12, "R"),
]


@router.get("/txt")
async def lina1334_txt(
    request: Request,
    fecha_hasta: str = Query(default=""),
):
    Lina1334.set_prog_code(PROG_CODE)
    user = Lina1334.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    try:
        fh = date.fromisoformat(fecha_hasta) if fecha_hasta else date.today()
    except ValueError:
        fh = date.today()

    empr_code, empr_name = Lina1334.get_empr_info()
    filas = _get_filas(fh)

    total_valor = sum(f[7] for f in filas)

    filas_txt = [
        [f[0], f[1], f[2], str(f[3]), str(f[4]), str(f[5]), fmt_money(f[6]), fmt_money(f[7])]
        for f in filas
    ]
    totales_txt = ["", "", "TOTAL", "", "", "", "", fmt_money(total_valor)]

    txt = generar_txt(
        titulo    = "Artículos Bajo Punto Mínimo de Pedido",
        subtitulo = f"Al {fh.strftime('%d/%m/%Y')}",
        columnas  = _COLUMNAS_TXT,
        filas     = filas_txt,
        app_name  = APP_CONFIG.get("app_name", ""),
        empr_code = empr_code,
        empr_name = empr_name,
        usuario   = user,
        fecha     = date.today().strftime("%d/%m/%Y"),
        hora      = datetime.now().strftime("%H:%M"),
        totales   = totales_txt,
    )

    return Response(
        content    = txt.encode("utf-8"),
        media_type = "text/plain; charset=utf-8",
        headers    = {"Content-Disposition": "inline; filename=articulos_bajo_pmp.txt"},
    )
