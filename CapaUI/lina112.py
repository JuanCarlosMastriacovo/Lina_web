from fastapi import APIRouter, Request, Query
from fastapi.responses import Response, HTMLResponse, RedirectResponse
from weasyprint import HTML
from jinja2 import Environment, FileSystemLoader
from pathlib import Path
from datetime import date, datetime
from io import BytesIO

import openpyxl

from CapaBRL.linabase import linabase
from CapaDAL.tablebase import get_table_model
from CapaDAL.config import APP_CONFIG
from CapaUI.xlsx_styles import (
    TITLE_FONT, SUBTITLE_FONT, HEADER_FONT, HEADER_FILL, HEADER_ALIGN,
)

# ==================== CONSTANTES Y ROUTER ====================

router     = APIRouter()
PROG_CODE  = "LINA112"
ROUTE_BASE = "/lina112"

LinaClie           = get_table_model("linaclie")
CLIENT_KEY_FIELD   = LinaClie.get_business_key_field()
CLIENT_LABEL_FIELD = LinaClie.get_selector_fields()[1]

pdf_templates_dir = Path(__file__).parent.parent / "templates"
pdf_jinja_env     = Environment(loader=FileSystemLoader(str(pdf_templates_dir)))


# ==================== CLASE PRINCIPAL ====================

class Lina112(linabase):
    """Módulo de listado de clientes (LINA112)."""
    pass


# ==================== FUNCIONES AUXILIARES ====================

def _get_clientes(conn, desde: int, hasta: int):
    """Obtiene clientes filtrados por rango de código."""
    clients = LinaClie.list_all(
        order_by=CLIENT_KEY_FIELD,
        fields=[CLIENT_KEY_FIELD, CLIENT_LABEL_FIELD],
        conn=conn,
    )
    return [
        [str(c.get(CLIENT_KEY_FIELD) or "0").zfill(4), str(c.get(CLIENT_LABEL_FIELD) or "")]
        for c in clients
        if desde <= int(c.get(CLIENT_KEY_FIELD) or 0) <= hasta
    ]


# ==================== RUTAS ====================

@router.get("/", response_class=HTMLResponse)
async def lina112_index(request: Request):
    Lina112.set_prog_code(PROG_CODE)
    user = Lina112.get_current_user(request)
    if not user:
        return RedirectResponse("/login")
    return Lina112.templates.TemplateResponse(
        "lina112/seleccion.html",
        {"request": request, "error": None},
    )


@router.get("/pdf")
async def lina112_pdf(
    request: Request,
    desde: int = Query(default=0),
    hasta: int = Query(default=9999),
):
    """Genera el PDF del listado de clientes."""
    Lina112.set_prog_code(PROG_CODE)
    user = Lina112.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    if desde > hasta:
        return Lina112.templates.TemplateResponse(
            "lina112/seleccion.html",
            {"request": request, "error": "El valor 'Desde' no puede ser mayor que 'Hasta'."},
        )

    empr_code, empr_name = Lina112.get_empr_info()
    conn  = Lina112.get_task_conn(request, readonly=True)
    filas = _get_clientes(conn, desde, hasta)

    template = pdf_jinja_env.get_template("lina112/main.html")
    html_str = template.render(
        app_name        = APP_CONFIG.get("app_name", ""),
        app_description = APP_CONFIG.get("app_description", ""),
        prog_code = PROG_CODE,
        empr_code = empr_code,
        empr_name = empr_name,
        usuario   = user,
        titulo    = "Listado de Clientes",
        subtitulo = f"Clientes {str(desde).zfill(4)} al {str(hasta).zfill(4)}",
        fecha     = date.today().strftime("%d/%m/%Y"),
        hora      = datetime.now().strftime("%H:%M"),
        columnas  = ["Código", "Nombre"],
        filas     = filas,
    )

    pdf = HTML(string=html_str).write_pdf()
    return Response(
        content    = pdf,
        media_type = "application/pdf",
        headers    = {"Content-Disposition": "inline; filename=clientes.pdf"},
    )


@router.get("/xlsx")
async def lina112_xlsx(
    request: Request,
    desde: int = Query(default=0),
    hasta: int = Query(default=9999),
):
    """Genera el XLSX del listado de clientes."""
    Lina112.set_prog_code(PROG_CODE)
    user = Lina112.get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    if desde > hasta:
        return Lina112.templates.TemplateResponse(
            "lina112/seleccion.html",
            {"request": request, "error": "El valor 'Desde' no puede ser mayor que 'Hasta'."},
        )

    empr_code, empr_name = Lina112.get_empr_info()
    conn  = Lina112.get_task_conn(request, readonly=True)
    filas = _get_clientes(conn, desde, hasta)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    header_font   = HEADER_FONT
    header_fill   = HEADER_FILL
    header_align  = HEADER_ALIGN
    title_font    = TITLE_FONT
    subtitle_font = SUBTITLE_FONT

    ws.merge_cells("A1:B1")
    ws["A1"]      = "Listado de Clientes"
    ws["A1"].font = title_font

    ws.merge_cells("A2:B2")
    ws["A2"]      = f"{APP_CONFIG.get('app_name', '')} — {empr_code} {empr_name}"
    ws["A2"].font = subtitle_font

    ws.merge_cells("A3:B3")
    ws["A3"]      = f"Clientes {str(desde).zfill(4)} al {str(hasta).zfill(4)} — Usuario: {user} — {date.today().strftime('%d/%m/%Y')} {datetime.now().strftime('%H:%M')}"
    ws["A3"].font = subtitle_font

    ws.append([])

    ws.append(["Código", "Nombre"])
    for cell in ws[5]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    for fila in filas:
        ws.append(fila)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 45

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        content    = buffer.read(),
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers    = {"Content-Disposition": "attachment; filename=clientes.xlsx"},
    )